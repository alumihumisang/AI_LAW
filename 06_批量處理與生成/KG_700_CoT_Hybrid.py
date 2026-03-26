#!/home/aru/AI_LAW/venv/bin/python3
"""
KG_700_CoT_Hybrid.py
混合模式：事實、法條和損害用標準方法，結論用CoT
優化版本：修正未成年誤判、改進損害項目處理、增強穩定性
"""

import os
import re
import json
import requests
import time
import sys
import argparse
from typing import List, Dict, Any, Optional
from collections import Counter

# 導入必要模組
try:
    import torch
    from transformers import AutoTokenizer, AutoModel
    from elasticsearch import Elasticsearch
    from neo4j import GraphDatabase
    from dotenv import load_dotenv
    FULL_MODE = True
    print("✅ 完整模式：所有檢索功能可用")
except ImportError as e:
    print(f"⚠️ 部分模組未安裝：{e}")
    print("⚠️ 使用簡化模式（僅LLM生成功能）")
    FULL_MODE = False

# 導入結構化金額處理器
try:
    from structured_legal_amount_processor import StructuredLegalAmountProcessor
    STRUCTURED_PROCESSOR_AVAILABLE = True
    print("✅ 結構化金額處理器載入成功")
except ImportError:
    STRUCTURED_PROCESSOR_AVAILABLE = False
    print("⚠️ 結構化金額處理器未找到")

# 導入基本金額標準化器
try:
    from legal_amount_standardizer import LegalAmountStandardizer
    BASIC_STANDARDIZER_AVAILABLE = True
    print("✅ 基本金額標準化器載入成功")
except ImportError:
    BASIC_STANDARDIZER_AVAILABLE = False
    print("⚠️ 基本金額標準化器未找到")

# ===== 基本設定 =====
LLM_URL = "http://localhost:11434/api/generate"
DEFAULT_MODEL = "gemma3:27b"

# ===== 檢索系統設定 =====
if FULL_MODE:
    # 載入環境變數
    env_path = os.path.join(os.path.dirname(__file__), '..', '01_設定與配置', '.env')
    load_dotenv(dotenv_path=env_path)
    
    # 嵌入模型設定
    BERT_MODEL = "shibing624/text2vec-base-chinese"
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    try:
        TOKENIZER = AutoTokenizer.from_pretrained(BERT_MODEL)
        MODEL = AutoModel.from_pretrained(BERT_MODEL, dtype=torch.float16 if torch.cuda.is_available() else torch.float32).to(device)
        print("✅ 嵌入模型載入成功")
    except Exception as e:
        print(f"❌ 嵌入模型載入失敗: {e}")
        FULL_MODE = False
    
    # ES 和 Neo4j 連接
    try:
        # 使用 requests 直接調用 ES API 避免版本兼容性問題
        ES_HOST = os.getenv("ELASTIC_HOST")
        ES_USER = os.getenv("ELASTIC_USER")
        ES_PASSWORD = os.getenv("ELASTIC_PASSWORD")
        ES_AUTH = (ES_USER, ES_PASSWORD)
        
        # 測試 ES 連接
        response = requests.get(f"{ES_HOST}/_cluster/health", auth=ES_AUTH, verify=False)
        if response.status_code != 200:
            raise Exception(f"ES連接失敗: {response.status_code}")
        
        NEO4J_DRIVER = GraphDatabase.driver(
            os.getenv("NEO4J_URI"),
            auth=(os.getenv("NEO4J_USERNAME"), os.getenv("NEO4J_PASSWORD")),
        )
        CHUNK_INDEX = "legal_kg_chunks"
        print("✅ 資料庫連接成功")
    except Exception as e:
        print(f"❌ 資料庫連接失敗: {e}")
        FULL_MODE = False
else:
    ES_HOST = None
    ES_AUTH = None
    NEO4J_DRIVER = None
    CHUNK_INDEX = None

# 案件類型對照表（ES檢索fallback用）
CASE_TYPE_MAP = {
    # 特殊案型如果找不到，fallback到相關基礎類型
    "§190動物案型": "單純原被告各一",
    "§188僱用人案型": "單純原被告各一", 
    "§187未成年案型": "單純原被告各一",
    
    # 複合案型的fallback
    "原被告皆數名+§188僱用人案型": "§188僱用人案型",
    "數名原告+§188僱用人案型": "§188僱用人案型",
    "數名被告+§188僱用人案型": "§188僱用人案型",
    "數名被告+§187未成年案型": "§187未成年案型", 
    "原被告皆數名+§187未成年案型": "§187未成年案型",
    "原被告皆數名+§190動物案型": "§190動物案型",
    
    # 基礎當事人數量類型（通常不需要fallback）
    "數名原告": "單純原被告各一",
    "數名被告": "單純原被告各一",
    "原被告皆數名": "單純原被告各一",
}

# ===== 輔助函數 =====
def extract_sections(text: str) -> dict:
    """提取文本段落"""
    result = {
        "accident_facts": "",
        "injuries": "",
        "compensation_facts": ""
    }
    
    # 事故發生緣由
    fact_match = re.search(r"一[、．.\s]*事故發生緣由[:：]?\s*(.*?)(?=二[、．.]|$)", text, re.S)
    if fact_match:
        result["accident_facts"] = fact_match.group(1).strip()
    
    # 受傷情形
    injury_match = re.search(r"二[、．.\s]*(?:原告)?受傷情形[:：]?\s*(.*?)(?=三[、．.]|$)", text, re.S)
    if injury_match:
        result["injuries"] = injury_match.group(1).strip()
    
    # 賠償事實根據
    comp_match = re.search(r"三[、．.\s]*請求賠償的事實根據[:：]?\s*(.*?)$", text, re.S)
    if comp_match:
        result["compensation_facts"] = comp_match.group(1).strip()
    
    return result

def extract_parties_with_llm(text: str) -> dict:
    """使用LLM提取當事人（更準確的方法）"""
    print("🤖 使用LLM智能提取當事人...")
    
    # 創建更精確的提示模板
    prompt = f"""請你幫我從以下車禍案件的法律文件中提取並列出所有原告和被告的真實姓名。

以下是案件內容：
{text}

🚨 **重要提取規則**：
1. ✅ **只能提取**明確標示為「原告○○○」的人
2. ❌ **絕對不能提取**標示為「訴外人○○○」的人（訴外人不是原告！）
3. ❌ **絕對不能提取**標示為「乘客」、「搭載」、「車上乘客」等非原告身份的人
4. ✅ 完整保留姓名，不可截斷（如：鄭凱祥不能寫成鄭祥）
5. ✅ 如果文中沒有明確的姓名，就直接寫「原告」、「被告」
6. ✅ 多個姓名用逗號分隔

輸出格式（只輸出這兩行）：
原告:姓名1,姓名2...
被告:姓名1,姓名2...

範例說明：
✅ 正確：
- 「原告吳麗娟」→ 原告:吳麗娟
- 「被告鄭凱祥」→ 被告:鄭凱祥
- 「原告陳皆宏駕駛...另一原告王惠華」→ 原告:陳皆宏,王惠華

❌ 錯誤：
- 「訴外人陳河田」→ 不是原告，絕對不提取！
- 「車上搭載乘客訴外人王惠滿」→ 不是原告，絕對不提取！
- 「搭載乘客李小明」→ 不是原告，絕對不提取！

📋 **判斷標準**：
- 必須在文中明確看到「原告XXX」才算是原告
- 只要看到「訴外人XXX」就絕對不是原告
- 只要看到「乘客XXX」或「搭載XXX」且沒有「原告」二字，就不是原告"""

    try:
        # 調用LLM
        response = requests.post(
            LLM_URL,
            json={
                "model": DEFAULT_MODEL,
                "prompt": prompt,
                "stream": False
            },
            timeout=60
        )
        
        if response.status_code == 200:
            llm_result = response.json()["response"].strip()
            print(f"🤖 LLM提取結果: {llm_result}")
            parties = parse_llm_parties_result(llm_result)
            # 驗證並修正可能被LLM截斷的姓名
            parties = verify_and_fix_party_names(parties, text)
            return parties
        else:
            print(f"❌ LLM調用失敗: {response.status_code}")
            return extract_parties_fallback(text)

    except Exception as e:
        print(f"❌ LLM提取異常: {e}")
        return extract_parties_fallback(text)

def parse_llm_parties_result(llm_result: str) -> dict:
    """解析LLM的當事人提取結果"""
    result = {"原告": "原告", "被告": "被告", "被告數量": 1, "原告數量": 1}

    # 檢查LLM是否返回了無效的回應
    invalid_responses = ["請提供", "無法提取", "沒有提供", "由於您沒有"]
    if any(invalid in llm_result for invalid in invalid_responses):
        print("⚠️ LLM返回無效回應，使用fallback")
        return result

    lines = llm_result.split('\n')

    for line in lines:
        line = line.strip()
        if line.startswith('原告:') or line.startswith('原告：'):
            plaintiff_text = line.split(':', 1)[1].strip() if ':' in line else line.split('：', 1)[1].strip()
            if plaintiff_text:
                # 分割多個原告
                plaintiffs = [p.strip() for p in plaintiff_text.split(',') if p.strip()]
                result["原告"] = "、".join(plaintiffs)
                result["原告數量"] = len(plaintiffs)

        elif line.startswith('被告:') or line.startswith('被告：'):
            defendant_text = line.split(':', 1)[1].strip() if ':' in line else line.split('：', 1)[1].strip()
            if defendant_text:
                # 分割多個被告
                defendants = [d.strip() for d in defendant_text.split(',') if d.strip()]
                # 過濾掉明顯無效的值（LLM 有時會把被告填成「原告」）
                invalid_defendant_values = {'原告', '被告', '原告本人', '不明', '未提及', ''}
                defendants = [d for d in defendants if d not in invalid_defendant_values]
                if defendants:
                    result["被告"] = "、".join(defendants)
                    result["被告數量"] = len(defendants)
                # 若過濾後為空，保留預設值「被告」

    return result

def verify_and_fix_party_names(parties: dict, original_text: str) -> dict:
    """驗證並修正LLM可能截斷的姓名"""
    import re

    # 從原文中用正則提取完整姓名作為參考
    reference_plaintiffs = set()
    reference_defendants = set()

    # 提取原告姓名（2-4個中文字）
    plaintiff_matches = re.findall(r'原告([\u4e00-\u9fff○]{2,4})', original_text)
    for name in plaintiff_matches:
        # 過濾掉明顯不是姓名的
        if name not in ['受有', '所有', '因本', '部分', '之損', '邱品']:
            reference_plaintiffs.add(name)

    # 提取被告姓名（2-4個中文字）
    defendant_matches = re.findall(r'被告([\u4e00-\u9fff○]{2,4})', original_text)
    for name in defendant_matches:
        if name not in ['受有', '所有', '因本', '部分', '應負']:
            reference_defendants.add(name)

    # 檢查並修正原告姓名
    if parties.get('原告') and parties['原告'] != '原告':
        llm_plaintiffs = [p.strip() for p in parties['原告'].replace('、', ',').split(',')]
        corrected_plaintiffs = []

        for llm_name in llm_plaintiffs:
            # 檢查這個姓名是否可能是截斷的
            possible_full_name = None

            if len(llm_name) == 2:
                # 兩字姓名，可能缺少中間字
                surname = llm_name[0]
                last_char = llm_name[1]

                # 在參考姓名中尋找完整版本
                for ref_name in reference_plaintiffs:
                    if len(ref_name) == 3 and ref_name[0] == surname and ref_name[2] == last_char:
                        possible_full_name = ref_name
                        print(f"⚠️ 修正截斷姓名：原告 '{llm_name}' → '{possible_full_name}'")
                        break

            corrected_plaintiffs.append(possible_full_name if possible_full_name else llm_name)

        parties['原告'] = '、'.join(corrected_plaintiffs)

    # 檢查並修正被告姓名（同樣邏輯）
    if parties.get('被告') and parties['被告'] != '被告':
        llm_defendants = [d.strip() for d in parties['被告'].replace('、', ',').split(',')]
        corrected_defendants = []

        for llm_name in llm_defendants:
            possible_full_name = None

            if len(llm_name) == 2:
                surname = llm_name[0]
                last_char = llm_name[1]

                for ref_name in reference_defendants:
                    if len(ref_name) == 3 and ref_name[0] == surname and ref_name[2] == last_char:
                        possible_full_name = ref_name
                        print(f"⚠️ 修正截斷姓名：被告 '{llm_name}' → '{possible_full_name}'")
                        break

            corrected_defendants.append(possible_full_name if possible_full_name else llm_name)

        parties['被告'] = '、'.join(corrected_defendants)

    return parties

def extract_parties_fallback(text: str) -> dict:
    """當LLM提取失敗時的fallback方法（簡化版正則）"""
    print("⚠️ 使用fallback方法提取當事人...")
    result = {"原告": "原告", "被告": "被告", "被告數量": 1, "原告數量": 1}
    
    # 簡化的正則表達式提取
    plaintiffs = set()
    defendants = set()
    
    # 基本模式
    plaintiff_patterns = [
        r'原告([\u4e00-\u9fff]{2,4})',
        r'原告([甲乙丙丁戊])'
    ]
    
    defendant_patterns = [
        r'被告([\u4e00-\u9fff]{2,4})',
        r'被告([甲乙丙丁戊])'
    ]
    
    for pattern in plaintiff_patterns:
        matches = re.findall(pattern, text)
        plaintiffs.update(matches)
    
    for pattern in defendant_patterns:
        matches = re.findall(pattern, text)
        defendants.update(matches)
    
    # 清理和組合結果
    if plaintiffs:
        result["原告"] = "、".join(sorted(plaintiffs))
        result["原告數量"] = len(plaintiffs)
    elif "原告" in text:
        result["原告"] = "原告"
    
    if defendants:
        result["被告"] = "、".join(sorted(defendants))
        result["被告數量"] = len(defendants)
    elif "被告" in text:
        result["被告"] = "被告"
    
    return result

def extract_parties(text: str) -> dict:
    """主要的當事人提取函數（優先使用LLM）"""
    return extract_parties_with_llm(text)

# ===== 檢索相關函數 =====
def embed(text: str):
    """文字向量化"""
    if not FULL_MODE:
        return []
    
    t = TOKENIZER(text, truncation=True, padding="max_length", max_length=512, return_tensors="pt")
    t = {k: v.to(device) for k, v in t.items()}
    with torch.no_grad():
        vec = MODEL(**t).last_hidden_state.mean(dim=1).squeeze()
    return vec.cpu().numpy().tolist()

def es_search(query_vector, case_type: str, top_k: int = 3, label: str = "Facts", quiet: bool = False):
    """ES 搜尋（含fallback機制）"""
    if not FULL_MODE or not ES_HOST:
        return []
    
    def _search(label_filter, case_type_filter):
        must_clause = [{"match": {"label": label_filter}}]
        
        if case_type_filter:
            # 嘗試多種可能的 case_type 欄位格式
            case_type_options = [
                {"term": {"case_type.keyword": case_type_filter}},
                {"term": {"case_type": case_type_filter}},
                {"match": {"case_type": case_type_filter}}
            ]
            
            # 使用 should 查詢，任一符合即可
            must_clause.append({
                "bool": {
                    "should": case_type_options,
                    "minimum_should_match": 1
                }
            })
            
        body = {
            "size": top_k,
            "query": {
                "script_score": {
                    "query": {"bool": {"must": must_clause}},
                    "script": {
                        "source": "cosineSimilarity(params.qv,'embedding')+1.0",
                        "params": {"qv": query_vector},
                    },
                }
            },
        }
        
        # 調試信息：只在非安靜模式下輸出
        if not quiet:
            print(f"🔍 ES查詢條件: index={CHUNK_INDEX}, label={label_filter}, case_type={case_type_filter}")
        
        try:
            url = f"{ES_HOST}/{CHUNK_INDEX}/_search"
            response = requests.post(url, auth=ES_AUTH, json=body, verify=False)
            if response.status_code == 200:
                result = response.json()
                hits = result["hits"]["hits"]
                total_docs = result["hits"]["total"]["value"] if isinstance(result["hits"]["total"], dict) else result["hits"]["total"]
                if not quiet:
                    print(f"📊 ES查詢結果: 找到 {len(hits)} 個匹配結果，總文檔數: {total_docs}")
                return hits
            else:
                if not quiet:
                    print(f"❌ ES查詢失敗: {response.status_code} - {response.text}")
                return []
        except Exception as e:
            if not quiet:
                print(f"❌ ES查詢失敗: {e}")
            return []

    if not quiet:
        print(f"🔎 使用 case_type='{case_type}' 搜索相似案例...")
    hits = _search(label, case_type)
    
    if not hits:
        # 先檢查索引映射和可用的案件類型
        try:
            # 檢查 mapping
            mapping_url = f"{ES_HOST}/{CHUNK_INDEX}/_mapping"
            mapping_response = requests.get(mapping_url, auth=ES_AUTH, verify=False)
            if mapping_response.status_code == 200:
                mapping = mapping_response.json()
                properties = mapping[CHUNK_INDEX]["mappings"]["properties"]
                has_case_type = "case_type" in properties
                print(f"🗺️ case_type欄位存在: {has_case_type}")
                
                if has_case_type:
                    # 嘗試不同的欄位名稱
                    for field_name in ["case_type", "case_type.keyword"]:
                        try:
                            check_body = {
                                "size": 0,
                                "aggs": {
                                    "case_type_count": {"terms": {"field": field_name, "size": 20}}
                                }
                            }
                            check_url = f"{ES_HOST}/{CHUNK_INDEX}/_search"
                            check_response = requests.post(check_url, auth=ES_AUTH, json=check_body, verify=False)
                            if check_response.status_code == 200:
                                check_result = check_response.json()
                                available_types = [bucket["key"] for bucket in check_result["aggregations"]["case_type_count"]["buckets"]]
                                print(f"📋 使用欄位 {field_name} 找到的案件類型: {available_types}")
                                break
                        except Exception as field_e:
                            print(f"⚠️ 欄位 {field_name} 查詢失敗: {field_e}")
                else:
                    print("❌ case_type欄位不存在於索引映射中")
            else:
                print(f"❌ 獲取映射失敗: {mapping_response.status_code}")
                
        except Exception as e:
            print(f"⚠️ 無法檢查索引映射或案件類型: {e}")
        
        fallback = CASE_TYPE_MAP.get(case_type, "單純原被告各一")
        if fallback != case_type:
            print(f"⚠️ 使用 fallback='{fallback}' 重新搜尋...")
            hits = _search(label, fallback)
    
    if not hits:
        print("⚠️ 不限案件類型進行搜尋...")
        hits = _search(label, None)
    
    return hits

def rerank_case_ids_by_paragraphs(query_text: str, case_ids: List[str], label: str = "Facts", quiet: bool = False) -> List[str]:
    """根據段落級資料重新排序案例"""
    if not FULL_MODE or not ES_HOST:
        return case_ids
    
    if not quiet:
        print("📘 啟動段落級 rerank...")
    try:
        from sklearn.metrics.pairwise import cosine_similarity
        import numpy as np
    except ImportError:
        if not quiet:
            print("⚠️ sklearn未安裝，跳過rerank")
        return case_ids

    query_vec = embed(query_text)
    if not query_vec:
        return case_ids
    
    query_vec_np = np.array(query_vec).reshape(1, -1)
    scored_cases = []
    
    for cid in case_ids:
        try:
            # 使用 requests 直接調用 ES API
            search_body = {
                "query": {
                    "bool": {
                        "must": [
                            {"term": {"case_id": cid}},
                            {"term": {"label": label}}
                        ]
                    }
                },
                "size": 1
            }
            url = f"{ES_HOST}/legal_kg_paragraphs/_search"
            response = requests.post(url, auth=ES_AUTH, json=search_body, verify=False)
            if response.status_code != 200:
                continue
            res = response.json()
            hits = res.get("hits", {}).get("hits", [])
            if hits:
                vec = hits[0]['_source']['embedding']
                para_vec_np = np.array(vec).reshape(1, -1)
                score = cosine_similarity(query_vec_np, para_vec_np)[0][0]
                scored_cases.append((cid, score))
        except Exception as e:
            print(f"⚠️ Case {cid} rerank失敗: {e}")
            scored_cases.append((cid, 0.0))

    scored_cases.sort(key=lambda x: x[1], reverse=True)
    return [cid for cid, _ in scored_cases]

def get_complete_cases_content(case_ids: List[str]) -> List[str]:
    """獲取完整案例內容"""
    if not FULL_MODE or not NEO4J_DRIVER:
        return []
    
    complete_cases = []
    
    try:
        with NEO4J_DRIVER.session() as session:
            for case_id in case_ids:
                # 查詢完整案例的事實段落
                result = session.run("""
                    MATCH (c:Case {case_id: $case_id})-[:包含]->(f:Facts)
                    RETURN f.description AS facts_content
                """, case_id=case_id).data()
                
                if result:
                    # 組合案例的所有事實段落
                    case_content = "\n".join([record["facts_content"] for record in result if record["facts_content"]])
                    if case_content:
                        complete_cases.append(case_content)
                else:
                    print(f"⚠️ 案例 {case_id} 無法獲取完整內容")
                    
    except Exception as e:
        print(f"⚠️ 獲取完整案例內容失敗: {e}")
    
    return complete_cases

def query_laws(case_ids):
    """從Neo4j查詢法條資訊"""
    if not FULL_MODE or not NEO4J_DRIVER:
        return Counter(), {}
    
    counter = Counter()
    law_text_map = {}
    
    try:
        with NEO4J_DRIVER.session() as session:
            for cid in case_ids:
                result = session.run("""
                    MATCH (c:Case {case_id: $cid})-[:包含]->(:Facts)-[:適用]->(l:Laws)-[:包含]->(ld:LawDetail)
                    RETURN collect(distinct ld.name) AS law_names, collect(distinct ld.text) AS law_texts
                """, cid=cid).single()
                if result:
                    names = result["law_names"]
                    texts = result["law_texts"]
                    counter.update(names)
                    for n, t in zip(names, texts):
                        if n not in law_text_map:
                            law_text_map[n] = t
    except Exception as e:
        print(f"⚠️ Neo4j查詢失敗: {e}")
    
    return counter, law_text_map

def get_similar_cases_laws_stats(case_ids):
    """獲取相似案例的法條統計資訊"""
    counter, _ = query_laws(case_ids)
    return counter.most_common()

def normalize_article_number(article: str) -> str:
    """條號格式標準化：第191-2條 → 第191條之2"""
    # 處理特殊格式的條號
    article = re.sub(r'第(\d+)-(\d+)條', r'第\1條之\2', article)
    return article

def detect_special_relationships(text: str, parties: dict) -> dict:
    """偵測特殊法律關係（優化版）"""
    # 更嚴格的被告數量判斷
    defendant_count = parties.get('被告數量', 0)
    if defendant_count == 0:
        # 如果沒有明確數量，嘗試從文本中檢測
        # 檢查是否有明確的多被告表述
        multi_defendant_patterns = [
            r'被告.*?、.*?被告',  # 被告A、被告B
            r'被告.*?及.*?被告',  # 被告A及被告B
            r'被告.*?與.*?被告',  # 被告A與被告B
            r'被告二人',
            r'被告三人',
            r'被告等人',
            r'共同.*?被告',
        ]
        has_multi_defendant = any(re.search(pattern, text) for pattern in multi_defendant_patterns)
        defendant_count = 2 if has_multi_defendant else 1

    relationships = {
        "未成年": False,
        "雇傭關係": False,
        "動物損害": False,
        "多被告": defendant_count > 1,
        "多原告": parties.get('原告數量', 1) > 1   # 新增多原告判斷
    }
    
    # 更精確的未成年檢測
    # 1. 明確提到未成年相關詞彙（但要排除子女、原告子女等情況）
    explicit_minor_keywords = ["法定代理人", "監護人", "未滿十八歲", "未滿18歲"]
    if any(keyword in text for keyword in explicit_minor_keywords):
        relationships["未成年"] = True

    # 檢查「未成年」但排除指涉子女的情況
    if "未成年" in text:
        # 排除：未成年子女、原告子女、扶養子女等
        exclude_patterns = [
            r'未成年子女',
            r'[一二三四五六七八九十數]名未成年',  # 二名未成年、三名未成年等
            r'原告.*?未成年',
            r'扶養.*?未成年',
            r'照顧.*?未成年',
            r'未成年.*?須.*?扶養',  # 未成年子女須扶養
            r'未成年.*?須.*?照顧',  # 未成年子女須照顧
        ]
        is_about_children = any(re.search(pattern, text) for pattern in exclude_patterns)
        if not is_about_children:
            # 只有在明確指涉被告是未成年時才判定
            if re.search(r'被告.*?未成年', text):
                relationships["未成年"] = True
    
    # 2. 檢查具體年齡（18歲以下）
    age_pattern = r'(\d+)\s*歲'
    age_matches = re.findall(age_pattern, text)
    for age_str in age_matches:
        age = int(age_str)
        if age < 18:
            relationships["未成年"] = True
            break
    
    # 3. 學校關鍵字需要更謹慎
    school_keywords = ["國中生", "國小生", "高中生"]  # 不是單純的"國中"、"高中"
    if any(keyword in text for keyword in school_keywords):
        relationships["未成年"] = True
    
    # 檢查雇傭關係（必須是被告與侵權人之間的雇傭關係）
    # 檢查是否有「被告僱用」或「執行職務」等明確表述
    employment_patterns = [
        # 最強信號：明確提到僱用人責任
        r'僱用人責任',
        r'雇用人責任',

        # 直接的僱傭關係描述（可能沒有「被告」字樣）
        r'受僱於.*?(?:被告|即)',  # 「受僱於被告」或「受僱於XX即YY」
        r'受雇於.*?(?:被告|即)',

        # 有「被告」字樣的模式
        r'被告.*?僱用',
        r'被告.*?雇主',
        r'被告.*?受僱',
        r'受僱.*?被告',
        r'受雇.*?被告',
        r'僱用.*?被告',

        # 執行職務相關
        r'執行.*?職務',  # 修改：支援「執行...職務」（中間可能有其他字）
        r'執行.*?業務',  # 新增：「執行拖吊業務」等
        r'職務上.*?行為',
        r'係在執行',

        # 其他僱傭關係暗示
        r'公司車',
        r'被告.*?員工',
        r'被告公司.*?員工',
        r'指揮.*?監督',
    ]
    # 排除：原告受僱於第三方公司的情況
    exclude_employment_patterns = [
        r'原告.*?受僱(?!.*被告)',  # 原告受僱但後面沒有提到被告
        r'原告.*?任職(?!.*被告)',
    ]

    # 檢查是否有僱傭關係(被告受僱於另一被告)
    has_defendant_employment = any(re.search(pattern, text) for pattern in employment_patterns)

    # 檢查是否僅有原告受僱(無被告僱傭關係)
    has_plaintiff_employment = any(re.search(pattern, text) for pattern in exclude_employment_patterns)

    # 只有當「被告有僱傭關係」時才為 True
    # 即使原告也有任職資訊(如「原告任職於XX公司」)也不影響
    relationships["雇傭關係"] = has_defendant_employment

    # Debug: 顯示僱傭關係偵測結果
    if has_defendant_employment:
        matched = [p for p in employment_patterns if re.search(p, text)]
        print(f"🔍 偵測到僱傭關係：匹配模式 {matched[:2]}...")  # 只顯示前2個
    
    # 檢查動物損害
    animal_keywords = ["狗", "貓", "犬", "動物", "寵物", "咬傷", "抓傷"]
    relationships["動物損害"] = any(keyword in text for keyword in animal_keywords)
    
    return relationships

def determine_case_type(accident_facts: str, parties: dict) -> str:
    """判斷案件類型（七大基本類型）"""
    relationships = detect_special_relationships(accident_facts, parties)
    
    # 優先判斷特殊法條類型（互斥優先級）
    if relationships["動物損害"]:
        return "§190動物案型"
    elif relationships["雇傭關係"]: 
        return "§188僱用人案型"
    elif relationships["未成年"]:
        return "§187未成年案型"
    
    # 其次判斷當事人數量類型
    elif relationships["多原告"] and relationships["多被告"]:
        return "原被告皆數名"
    elif relationships["多原告"]:
        return "數名原告"
    elif relationships["多被告"]:
        return "數名被告"
    else:
        return "單純原被告各一"

def determine_applicable_laws(accident_facts: str, injuries: str, comp_facts: str, parties: dict) -> List[str]:
    """根據案件事實智能判斷適用法條"""
    applicable_laws = []

    # 偵測特殊關係
    relationships = detect_special_relationships(accident_facts + injuries + comp_facts, parties)

    # Debug: 顯示偵測結果
    print(f"🔍 法律關係偵測：{relationships}")
    
    # 1. 第184條第1項前段 - 基本侵權責任（必須）
    applicable_laws.append("民法第184條第1項前段")

    # 2. 車禍案件 - 第191條之2（交通工具）
    # ⚠️ 重要：動物案型不適用191-2（動物不是車輛）
    if not relationships["動物損害"]:
        traffic_keywords = ["汽車", "機車", "車輛", "駕駛", "交通", "撞", "碰撞"]
        if any(keyword in accident_facts for keyword in traffic_keywords):
            applicable_laws.append("民法第191條之2")
    
    # 3. 身體健康損害 - 第193條第1項
    health_damage_keywords = ["醫療", "看護", "工作損失", "薪資", "收入", "勞動能力"]
    if injuries or any(keyword in comp_facts for keyword in health_damage_keywords):
        applicable_laws.append("民法第193條第1項")
    
    # 4. 精神慰撫金 - 第195條第1項前段
    mental_damage_keywords = ["精神", "慰撫", "痛苦", "名譽", "人格"]
    if any(keyword in comp_facts for keyword in mental_damage_keywords):
        applicable_laws.append("民法第195條第1項前段")
    
    # 5. 特殊情況處理
    # 187條（未成年）與188條（僱傭）、185條（共同侵權）互斥
    # 188條（僱傭）與185條（共同侵權）互斥（僱傭關係不是共同侵權）
    if relationships["未成年"]:
        # 優先：未成年案件 - 第187條第1項
        applicable_laws.append("民法第187條第1項")
    elif relationships["雇傭關係"]:
        # 次優先：僱傭關係 - 第188條第1項本文
        # 注意：僱用人責任不適用185條（僱用人本身沒有侵權行為）
        applicable_laws.append("民法第188條第1項本文")
    elif relationships["多被告"]:
        # 多被告共同侵權 - 第185條第1項（真正的多人共同參與侵權）
        applicable_laws.append("民法第185條第1項")
    
    # 7. 動物損害 - 第190條第1項
    if relationships["動物損害"]:
        applicable_laws.append("民法第190條第1項")
    
    # 標準化條號格式
    applicable_laws = [normalize_article_number(law) for law in applicable_laws]
    
    return list(dict.fromkeys(applicable_laws))  # 去重但保持順序


    

# ===== 混合模式生成器 =====
class HybridCoTGenerator:
    """混合模式生成器：事實法條損害用標準，結論用CoT"""
    
    def __init__(self, model_name: str = DEFAULT_MODEL):
        self.model_name = model_name
        self.llm_url = LLM_URL
        
        # 初始化金額處理器
        if STRUCTURED_PROCESSOR_AVAILABLE:
            self.structured_processor = StructuredLegalAmountProcessor()
        else:
            self.structured_processor = None
        
        if BASIC_STANDARDIZER_AVAILABLE:
            self.basic_standardizer = LegalAmountStandardizer()
        else:
            self.basic_standardizer = None
        
        # 檢查LLM連接
        self.llm_available = self._check_llm_connection()
    
    def _check_llm_connection(self) -> bool:
        """檢查LLM連接"""
        try:
            response = requests.get("http://localhost:11434/api/version", timeout=5)
            return response.status_code == 200
        except:
            return False
    
    def call_llm(self, prompt: str, timeout: int = 180) -> str:
        """調用LLM"""
        if not self.llm_available:
            return "❌ LLM服務不可用"
        
        try:
            response = requests.post(
                self.llm_url,
                json={
                    "model": self.model_name,
                    "prompt": prompt,
                    "stream": False
                },
                timeout=timeout
            )
            
            if response.status_code == 200:
                result = response.json()["response"]
                # Ensure result is always a string
                if isinstance(result, dict):
                    print(f"⚠️ LLM returned dict instead of string: {result}")
                    result = str(result)
                return result.strip() if isinstance(result, str) else str(result).strip()
            else:
                return f"❌ LLM API錯誤: {response.status_code}"
                
        except Exception as e:
            return f"❌ LLM調用失敗: {str(e)}"
    
    def _chinese_num(self, num: int) -> str:
        """數字轉中文（支持到99）"""
        if num <= 0 or num > 99:
            return str(num)

        chinese_digits = ["", "一", "二", "三", "四", "五", "六", "七", "八", "九"]

        if num <= 10:
            if num == 10:
                return "十"
            return chinese_digits[num]
        elif num < 20:
            # 11-19: 十一、十二...十九
            return "十" + chinese_digits[num - 10]
        else:
            # 20-99: 二十、二十一...九十九
            tens = num // 10
            ones = num % 10
            result = chinese_digits[tens] + "十"
            if ones > 0:
                result += chinese_digits[ones]
            return result
    
    def _extract_all_plaintiffs(self, text: str) -> List[str]:
        """提取所有原告姓名"""
        plaintiffs = []
        
        # 方法1：從文本中找「原告○○○」的模式
        pattern = r'原告([^，。；、\s]{2,4})(?:為|因|受|前往|支出)'
        matches = re.findall(pattern, text)
        plaintiffs.extend(matches)
        
        # 去重並保持順序
        seen = set()
        unique_plaintiffs = []
        for p in plaintiffs:
            if p not in seen:
                seen.add(p)
                unique_plaintiffs.append(p)
        
        return unique_plaintiffs
    
    def generate_standard_facts(self, accident_facts: str, similar_cases: List[str] = None) -> str:
        """標準方式生成事實段落（完整保留律師輸入）"""
        print("📝 轉換事實段落格式（保持原文完整）...")

        # 注意：不使用相似案例參考，避免LLM被誤導去改寫

        prompt = f"""你是台灣律師助理，請將律師提供的事故經過內容轉換為起訴狀格式。

⚠️ **核心原則**：你的工作是「格式轉換」，不是「內容改寫」！

律師提供的事故經過（原文）：
───────────────────────────
{accident_facts}
───────────────────────────

📋 **唯一任務**：
1. 移除開頭的標題（如「一、事故發生緣由:」或類似標題）
2. 在內容開頭加上「一、緣」兩個字
3. **其餘內容100%保持原樣**

✅ **必須保留**（完整不變）：
- 所有姓名（如「被告乙○○」、「被告甲○○」、「原告」）
- 所有時間（如「民國111年2月10日上午11時35分許」）
- 所有地點（如「臺南市永康區中華路與勝華街之路口處」）
- 所有車輛資訊（如「車牌號碼000-0000號普通重型機車」、「下稱A車」、「下稱B車」、「下稱C車」）
- 所有法律用語（如「依其情形無不能注意之情事」、「竟疏未注意」、「適有」、「致」）
- 所有事故細節描述（完整保留，一字不改）
- 原文的段落結構（如果原文是一整段，輸出也必須是一整段）

❌ **嚴格禁止**：
- ❌ 改寫任何句子或用詞
- ❌ 簡化或省略任何內容
- ❌ 改變事故的敘述順序或方式
- ❌ 拆分成多個段落（如「一、」「二、」「三、」）
- ❌ 刪除「A車」、「B車」、「C車」等代稱
- ❌ 刪除「下稱」、「適有」等連接詞
- ❌ 調整或優化任何語句

🎯 **正確範例**：

【輸入】
一、事故發生緣由:
被告於民國110年5月1日下午3時許，駕駛車輛沿台北市XX路行駛，竟疏未注意車前狀況，撞及前方原告車輛。

【正確輸出】
一、緣被告於民國110年5月1日下午3時許，駕駛車輛沿台北市XX路行駛，竟疏未注意車前狀況，撞及前方原告車輛。

⚠️ **最後提醒**：請完整複製律師提供的內容，只做最小限度的格式調整！

請直接輸出轉換後的事實段落（不要包含任何說明或註解）："""
        
        result = self.call_llm(prompt)

        # 清理括號提醒文字
        result = self._remove_bracket_reminders(result)

        # 清理LLM可能加上的說明文字
        result = result.strip()

        # 提取事實段落（改用貪婪匹配，確保獲取完整內容）
        # 匹配「一、」開頭直到文本結尾，保留完整內容
        fact_match = re.search(r"一、\s*(.*)", result, re.S)
        if fact_match:
            cleaned_content = fact_match.group(1).strip()
            return f"一、{cleaned_content}"

        # 如果沒有「一、」開頭，但有「緣被告」，直接使用
        if "緣被告" in result or "緣原告" in result:
            # 去除可能的多餘換行，但保留段落內的換行
            cleaned_result = result.strip()
            # 如果已經有「一、」就直接返回
            if cleaned_result.startswith("一、"):
                return cleaned_result
            # 否則加上「一、」
            return f"一、{cleaned_result}"

        # Fallback：直接從原文轉換（移除標題，加上「一、緣」）
        # 移除可能的標題行（如「一、事故發生緣由:」）
        facts_content = re.sub(r'^一、事故發生緣由[:：]\s*', '', accident_facts.strip())
        # 如果內容已經有「緣」開頭，就不重複加
        if facts_content.startswith("緣"):
            return f"一、{facts_content}"
        else:
            return f"一、緣{facts_content}"
    
    def generate_standard_laws(self, accident_facts: str, injuries: str, parties: dict, compensation_facts: str = "") -> str:
        """標準方式生成法律依據（符合法條引用規範）"""
        print("⚖️ 使用標準方式生成法律依據...")
        
        # 智能判斷適用法條
        applicable_laws = determine_applicable_laws(accident_facts, injuries, compensation_facts, parties)
        
        # 完整的法條說明對照表（精確到項、段、但書）
        law_descriptions = {
            "民法第184條第1項前段": "因故意或過失，不法侵害他人之權利者，負損害賠償責任。",
            "民法第185條第1項": "數人共同不法侵害他人之權利者，連帶負損害賠償責任。",
            "民法第187條第1項": "無行為能力人或限制行為能力人，不法侵害他人之權利者，以行為時有識別能力為限，與其法定代理人連帶負損害賠償責任。",
            "民法第188條第1項本文": "受僱人因執行職務，不法侵害他人之權利者，由僱用人與行為人連帶負損害賠償責任。",
            "民法第190條第1項": "動物加損害於他人者，由其占有人負損害賠償責任。",
            "民法第191條之2": "汽車、機車或其他非依軌道行駛之動力車輛，在使用中加損害於他人者，駕駛人應賠償因此所生之損害。",
            "民法第193條第1項": "不法侵害他人之身體或健康者，對於被害人因此喪失或減少勞動能力或增加生活上之需要時，應負損害賠償責任。",
            "民法第195條第1項前段": "不法侵害他人之身體、健康、名譽、自由、信用、隱私、貞操，或不法侵害其他人格法益而情節重大者，被害人雖非財產上之損害，亦得請求賠償相當之金額。"
        }
        
        # 組合法條內容（先列法條內容）
        law_texts = []
        valid_laws = []
        
        for law in applicable_laws:
            if law in law_descriptions:
                law_texts.append(f"「{law_descriptions[law]}」")
                valid_laws.append(law)
        
        if not law_texts:
            # Fallback：至少包含基本侵權條文
            law_texts = ["「因故意或過失，不法侵害他人之權利者，負損害賠償責任。」"]
            valid_laws = ["民法第184條第1項前段"]
        
        # 按正確格式組合：先法條內容，後條號
        law_content_block = "、".join(law_texts)
        article_list = "、".join(valid_laws)
        
        print(f"✅ 適用法條: {', '.join(valid_laws)}")
        
        return f"""二、按{law_content_block}，{article_list}分別定有明文。查被告因上開侵權行為，致原告受有下列損害，依前揭規定，被告應負損害賠償責任："""
    
    def _parse_damage_from_sentence(self, sentence: str, plaintiff: str) -> List[dict]:
        """從句子中解析損害項目（改進版）"""
        damages = []
        
        # 擴充損害類型模式
        damage_patterns = [
            {
                'keywords': ['醫療費用', '治療', '醫院', '診所', '就診'],
                'name': '醫療費用',
                'template': f'原告{plaintiff}因本次事故受傷，前往醫院治療所支出之醫療費用'
            },
            {
                'keywords': ['交通費'],
                'name': '交通費用',
                'template': f'原告{plaintiff}因本次事故所生之交通費用'
            },
            {
                'keywords': ['工資損失', '不能工作', '工作損失', '無法工作', '休養'],
                'name': '工作損失',
                'template': f'原告{plaintiff}因本次事故無法工作之收入損失'
            },
            {
                'keywords': ['慰撫金', '精神'],
                'name': '精神慰撫金',
                'template': f'原告{plaintiff}因本次事故所受精神上痛苦之慰撫金'
            },
            {
                'keywords': ['車輛貶值', '貶損', '價值減損', '交易價值'],
                'name': '車輛貶值損失',
                'template': f'系爭車輛因本次事故貶值之損失'
            },
            {
                'keywords': ['鑑定費'],
                'name': '鑑定費用',
                'template': f'為評估車輛損失所支出之鑑定費用'
            }
        ]
        
        # 更精確的金額提取（考慮前後文）
        # 不要只看到金額就提取，要看金額前後的描述
        
        return damages
    
    def generate_smart_compensation(self, injuries: str, comp_facts: str, parties: dict) -> str:
        """智能生成損害項目（簡化版本）"""
        print("💰 生成損害賠償...")
        
        # 直接使用LLM處理，避免複雜的結構化處理
        return self._generate_llm_based_compensation(comp_facts, parties)

    def _preprocess_chinese_numbers(self, text: str) -> str:
        """預處理中文數字，轉換為阿拉伯數字"""
        import re
        
        # 處理 X萬Y,YYY元 格式 (如：26萬4,379元)
        pattern1 = r'(\d+)萬(\d+,?\d+)元'
        def replace1(match):
            wan = int(match.group(1))
            rest = int(match.group(2).replace(',', ''))
            total = wan * 10000 + rest
            return f"{total:,}元"
        text = re.sub(pattern1, replace1, text)
        
        # 處理 X萬Y千元 格式 (如：30萬5千元)
        pattern2 = r'(\d+)萬(\d+)千元'
        def replace2(match):
            wan = int(match.group(1))
            qian = int(match.group(2))
            total = wan * 10000 + qian * 1000
            return f"{total:,}元"
        text = re.sub(pattern2, replace2, text)
        
        # 處理 X萬元 格式 (如：20萬元)
        pattern3 = r'(\d+)萬元'
        def replace3(match):
            wan = int(match.group(1))
            total = wan * 10000
            return f"{total:,}元"
        text = re.sub(pattern3, replace3, text)
        
        # 處理 X千元 格式 (如：5千元)
        pattern4 = r'(\d+)千元'
        def replace4(match):
            qian = int(match.group(1))
            total = qian * 1000
            return f"{total:,}元"
        text = re.sub(pattern4, replace4, text)
        
        return text
    
    def _remove_bracket_reminders(self, text: str) -> str:
        """移除文本中的括號提醒文字"""
        import re
        
        # 移除各種括號提醒模式
        patterns = [
            r'\（[^）]*請填寫[^）]*\）',  # （姓名：請填寫...）
            r'\（[^）]*請[^）]*\）',     # （請...）
            r'\（[^）]*：[^）]*\）',     # （任何：說明）
            r'\（[^）]*填寫[^）]*\）',   # （...填寫...）
            r'\（[^）]*輸入[^）]*\）',   # （...輸入...）
            r'\（[^）]*補充[^）]*\）',   # （...補充...）
        ]
        
        cleaned_text = text
        for pattern in patterns:
            cleaned_text = re.sub(pattern, '', cleaned_text)
        
        # 清理多餘的空格
        cleaned_text = re.sub(r'\s+', ' ', cleaned_text).strip()
        
        return cleaned_text
    
    def _remove_conclusion_phrases(self, text: str) -> str:
        """移除文本中的結論性文字和總計說明"""
        import re
        
        # 分行處理
        lines = text.split('\n')
        cleaned_lines = []
        
        for line in lines:
            line = line.strip()
            
            # 跳過包含結論性關鍵詞的行
            conclusion_keywords = [
                '綜上所述', '總計', '合計', '共計', '法定利息', 
                '按週年利率', '按年息', '起訴狀繕本送達',
                '清償日止', '自.*起至.*止', '利息', '總額',
                '此有相關收據可證', '有收據為證', '有統一發票可證', '可證',
                '經查', '查明', '經審理'
            ]
            
            should_skip = False
            for keyword in conclusion_keywords:
                if keyword in line:
                    should_skip = True
                    break
            
            if not should_skip and line:  # 保留非空且非結論性的行
                cleaned_lines.append(line)
        
        # 重新組合
        cleaned_text = '\n'.join(cleaned_lines)
        
        # 額外清理：移除可能的結論段落和證據文字
        # 移除從「綜上」開始到結尾的所有內容
        cleaned_text = re.sub(r'綜上.*$', '', cleaned_text, flags=re.MULTILINE | re.DOTALL)
        
        # 移除句子中的證據相關文字
        evidence_patterns = [
            r'，此有[^。]*可證。?',
            r'，有[^。]*收據[^。]*證。?',
            r'，有[^。]*發票[^。]*證。?',
            r'，[^。]*為證。?'
        ]
        
        for pattern in evidence_patterns:
            cleaned_text = re.sub(pattern, '。', cleaned_text)
        
        # 清理多餘的句號
        cleaned_text = re.sub(r'。+', '。', cleaned_text)
        
        return cleaned_text.strip()
    
    def _detect_structure_type(self, text: str) -> str:
        """檢測文本結構類型"""
        # 檢查結構化模式
        structured_patterns = [
            r'\d+\.\s*[^：:]+[：:]\s*[0-9,]+元',  # 1. 項目：金額
            r'（[一二三四五六七八九十]+）.*?[：:]\s*[0-9,]+元',  # （一）項目：金額
            r'說明：.*?[0-9,]+元'  # 說明：...金額
        ]
        
        structured_matches = sum(1 for pattern in structured_patterns 
                               if re.search(pattern, text))
        
        if structured_matches >= 2:
            return "structured"
        elif "元" in text:
            return "semi_structured"
        else:
            return "unstructured"
    
    def _generate_structured_compensation(self, comp_facts: str, parties: dict) -> str:
        """使用結構化處理器生成損害項目"""
        print("🏗️ 使用結構化處理器分析損害項目...")
        
        # 使用結構化處理器
        result = self.structured_processor.process_structured_document(comp_facts)
        
        # 檢查是否有計算錯誤
        validation = result.get('validation', {})
        if validation.get('claimed_total') and not validation.get('match', True):
            print(f"⚠️ 發現計算錯誤：原起訴狀聲稱{validation['claimed_total']:,}元，實際應為{validation['calculated_total']:,}元")
            print(f"📊 差額：{abs(validation['difference']):,}元")
        
        # 生成修正後的損害項目
        structured_text = ""
        
        # 按原告分組顯示
        current_plaintiff = None
        plaintiff_index = 0
        item_counter = 1
        
        for item in result['structured_items']:
            # 判斷是否為新的原告
            item_title = item['item_title']
            if '原告' in item_title:
                # 提取原告姓名
                plaintiff_match = re.search(r'原告([^之的]+)', item_title)
                if plaintiff_match:
                    plaintiff_name = plaintiff_match.group(1).strip()
                    if plaintiff_name != current_plaintiff:
                        current_plaintiff = plaintiff_name
                        plaintiff_index += 1
                        chinese_num = self._chinese_num(plaintiff_index)
                        structured_text += f"（{chinese_num}）原告{current_plaintiff}之損害：\n"
                        item_counter = 1
            
            # 添加損害項目
            structured_text += f"{item_counter}. {item['item_title']}：{item['formatted_amount']}\n"
            if item['description']:
                structured_text += f"   {item['description']}\n"
            item_counter += 1
            structured_text += "\n"
        
        # 添加我說明
        total_amount = result['calculation']['total']
        structured_text += f"\n💰 損害總計：新台幣{total_amount:,}元整\n"
        
        # 如果有計算錯誤，添加說明
        if validation.get('claimed_total') and not validation.get('match', True):
            structured_text += f"\n（註：經重新計算，正確總額為{total_amount:,}元，"
            if validation['difference'] > 0:
                structured_text += f"原起訴狀少算{validation['difference']:,}元）"
            else:
                structured_text += f"原起訴狀多算{abs(validation['difference']):,}元）"
        
        return structured_text
    
    def generate_cot_conclusion_with_smart_amount_calculation(self, accident_facts: str, compensation_text: str, parties: dict, damage_section: str = "") -> str:
        """使用智能金額計算生成CoT結論（防止重複和錯誤計算）"""
        print("🧠 生成CoT結論（含總金額計算）...")

        # 優先從 damage_section 提取金額（已結構化），否則從原始輸入提取
        if damage_section:
            amounts = self._extract_amounts_from_damage_section(damage_section)
            print(f"💰 從損害項目段落提取金額: {amounts}")
        else:
            amounts = self._extract_valid_claim_amounts(compensation_text)
            print(f"💰 從原始輸入提取金額: {amounts}")

        total_amount = sum(amounts) if amounts else 0
        print(f"💰 計算總額: {total_amount:,}元")

        # 計算損害項目的編號（從damage_section中提取）
        damage_count = 0
        if damage_section:
            # 計算（一）（二）（三）等的數量
            damage_count = len(re.findall(r'[（(][一二三四五六七八九十百]+[）)]', damage_section))

        # 構建包含金額計算的CoT提示詞
        plaintiff = parties.get("原告", "原告")
        defendant = parties.get("被告", "被告")

        # 生成下一個編號
        next_number_chinese = self._chinese_num(damage_count + 1) if damage_count > 0 else "一"

        # 檢查是否為多原告案件（使用parties中的原告數量，更準確）
        plaintiff_count = parties.get('原告數量', 1)
        is_multi_plaintiff = plaintiff_count > 1

        # 檢查是否為多被告案件
        defendant_count = parties.get('被告數量', 1)
        is_multi_defendant = defendant_count > 1

        prompt = f"""你是台灣資深律師，請運用Chain of Thought推理方式生成專業的起訴狀結論段落。

👥 當事人資訊：
原告：{plaintiff}（共{parties.get('原告數量', 1)}名）
被告：{defendant}（共{defendant_count}名）

📄 案件事實：
{accident_facts}

📄 已整理的損害項目段落：
{damage_section if damage_section else compensation_text}

💰 智能金額分析結果：
提取到的有效求償金額：{amounts}
正確總計：{total_amount:,}元

🧠 請使用Chain of Thought方式分析：

步驟1: 分析案件性質和當事人責任
步驟2: 從「已整理的損害項目段落」中識別各項損害項目和金額
步驟3: {"如果有多位原告，請依照原告分組整理各自的損害項目" if is_multi_plaintiff else "整理所有損害項目"}
步驟4: 形成簡潔精確的結論

🏛️ 最後請生成專業的結論段落，格式要求：

📋 **標準格式（請嚴格遵守）**："""

        if is_multi_plaintiff:
            # 多原告案件（無論被告數量）
            liability_term = "連帶賠償" if is_multi_defendant else "賠償"
            prompt += f"""
「（{next_number_chinese}）綜上所陳，被告應{liability_term}原告之損害，就原告[姓名1]部分，包含[項目1名稱][金額1]元、[項目2名稱][金額2]元...；就原告[姓名2]部分，包含[項目1名稱][金額1]元、[項目2名稱][金額2]元...，總計{total_amount:,}元，並自起訴狀副本送達翌日起至清償日止，按年息5%計算之利息。」

📝 **多原告特別要求**：
- ✅ 必須按原告分組：「就原告XXX部分，包含...；就原告YYY部分，包含...」
- ⚠️ **重要**：使用「已整理的損害項目段落」中實際出現的原告姓名，不要使用範例中的姓名
- ⚠️ **姓名完整性**：必須完整保留原告的全名，不可遺漏任何字（例如：「羅靖崴」不能寫成「羅崴」，「陳皆宏」不能寫成「陳宏」）
- ✅ 每位原告的項目用頓號「、」連接
- ✅ 原告之間用分號「；」分隔
- ✅ 項目名稱要簡潔：醫療費用、看護費用、薪資損失、精神慰撫金等
- {'✅ 多被告案件使用「連帶賠償」而非「賠償」' if is_multi_defendant else ''}
"""
        else:
            # 單一原告案件
            liability_term = "連帶賠償" if is_multi_defendant else "賠償"
            prompt += f"""
「（{next_number_chinese}）綜上所陳，被告應{liability_term}原告之損害，包含[項目1名稱][金額1]元、[項目2名稱][金額2]元、[項目3名稱][金額3]元...，總計{total_amount:,}元，並自起訴狀副本送達翌日起至清償日止，按年息5%計算之利息。」

📝 **單一原告案件特別要求**：
- ⚠️ **禁止使用**「就原告XXX部分」的格式（這是多原告案件才用的格式！）
- ✅ 正確格式：「被告應{liability_term}原告之損害，包含醫療費用XXX元、看護費用XXX元...」
- ✅ 只需列出「項目名稱」和「金額」，如：醫療費用54,741元、看護費用74,000元
- ✅ 單一原告不需要分組，不需要提及原告姓名，直接列出所有損害項目即可
- {'✅ **重要**：多被告案件必須使用「被告應連帶賠償」，不是「被告應賠償」' if is_multi_defendant else ''}
"""

        prompt += """
- ❌ 不要加入事故經過描述（如：被告因未遵守交通規則...）
- ❌ 不要加入責任論述（如：被告應負侵權行為責任...）
- ❌ 不要重複說明同一項目
- ✅ 最後以「總計X元，並自起訴狀副本送達翌日起至清償日止，按年息5%計算之利息。」結尾

⚠️ **金額格式**：
- ✅ 正確：20,900元、66,631元、180,000元、1,766,404元
- ❌ 錯誤：2萬0,900元、新台幣20900元

請直接輸出結論段落（不要包含任何推理過程）："""

        result = self.call_llm(prompt, timeout=180)

        # 後處理：修正混合格式的金額
        if result:
            result = self._fix_mixed_amount_format(result)
            # 修正可能被LLM截斷的原告姓名
            result = self._fix_plaintiff_names(result, parties)

        return result if result else f"（{next_number_chinese}）綜上所陳\n（LLM生成失敗，請檢查輸入內容）"

    def _fix_mixed_amount_format(self, text: str) -> str:
        """修正混合格式的金額表示（如：2萬0,900 → 20,900、99萬 → 990,000）"""
        import re

        def convert_with_remainder(match):
            """轉換有餘數的混合格式（如：2萬0,900）"""
            wan = int(match.group(1))
            rest = match.group(2).replace(',', '')
            total = wan * 10000 + int(rest)
            return f"{total:,}"

        def convert_without_remainder(match):
            """轉換純萬的格式（如：99萬）"""
            wan = int(match.group(1))
            total = wan * 10000
            return f"{total:,}"

        # 先處理有餘數的（如：2萬0,900）
        text = re.sub(r'(\d+)萬(\d+(?:,\d{3})*)', convert_with_remainder, text)

        # 再處理純萬的（如：99萬），但要確保後面是"元"或結尾
        text = re.sub(r'(\d+)萬(?=元|$|，|、)', convert_without_remainder, text)

        # 處理沒有千分位逗號的純阿拉伯數字（如：53122元 → 53,122元）
        def add_thousand_separator(match):
            """為純數字添加千分位逗號"""
            num_str = match.group(1)
            # 如果已經有逗號就不處理
            if ',' in num_str:
                return match.group(0)
            # 轉成整數再格式化
            num = int(num_str)
            return f"{num:,}元"

        # 匹配「純數字+元」的模式,但排除已經有逗號的
        # 使用負向前瞻確保前面不是逗號或數字
        text = re.sub(r'(?<![,\d])(\d{4,})元', add_thousand_separator, text)

        return text

    def generate_cot_conclusion_with_structured_analysis(self, accident_facts: str, compensation_text: str, parties: dict, damage_section: str = "") -> str:
        """使用智能金額計算生成CoT結論（簡化版本）"""
        print("🧠 生成CoT結論（使用智能金額計算）...")

        # 直接使用智能金額計算方式，避免複雜的結構化處理
        return self.generate_cot_conclusion_with_smart_amount_calculation(
            accident_facts, compensation_text, parties, damage_section
        )
    
    def generate_dual_round_cot_lawsuit(self, accident_facts: str, compensation_text: str, parties: dict, similar_cases: list = None) -> str:
        """雙輪CoT起訴狀生成：輪1生成，輪2檢查驗證"""
        print("🔄 啟動雙輪CoT起訴狀生成系統...")
        
        # 輪1：生成起訴狀
        print("📝 輪1：生成起訴狀...")
        round1_result = self._round1_generate_lawsuit(accident_facts, compensation_text, parties, similar_cases)
        
        # 輪2：檢查驗證
        print("🔍 輪2：檢查驗證...")
        round2_result = self._round2_verify_lawsuit(round1_result, accident_facts, compensation_text, parties)
        
        return round2_result
    
    def _round1_generate_lawsuit(self, accident_facts: str, compensation_text: str, parties: dict, similar_cases: list = None) -> str:
        """輪1：生成起訴狀 Prompt"""
        plaintiff = parties.get("原告", "原告")
        defendant = parties.get("被告", "被告")
        
        # 分析當事人狀況以確定法條引用
        case_analysis = self._analyze_case_for_laws(parties, accident_facts, compensation_text)
        
        prompt = f"""你是一名法律文書生成助手，負責撰寫交通事故民事賠償案件的四段式起訴狀。
請嚴格遵守以下規則：

👥 當事人資訊：
原告：{plaintiff}
被告：{defendant}

📄 案件事實：
{accident_facts}

📄 損害賠償內容：
{compensation_text}

一、段落結構：
1. 一、僅描述事故經過，不帶入傷勢或金額
2. 二、引用適用條文並說明構成要件
3. 三、詳列各項損害及金額（從第一個損害項目的左括號開始）
4. 綜上所陳：綜合請求賠償總額

二、法條引用條件（判斷→引用）：

權利受損 → 民法184條1項前段

交通工具加害 → 民法191條之2

精神慰撫金/非財產損害 → 民法195條1項前段

工作損失/薪資損失/看護費用 → 民法193條1項

被告為未成年人且有法定代理人成為被告 → 必須引用民法187條1項，且不得引用185條

僱用關係（受僱人+雇主） → 必須引用民法188條1項本文，且不得引用185條

僅當未符合187或188，且存在多名被告 → 引用185條1項

動物造成損害 → 民法190條1項

三、格式：
條文引用格式為「文字先引用，再列條號」。

🔍 案件分析結果：
{case_analysis}

請嚴格按照上述規則生成完整的四段式起訴狀："""

        result = self.call_llm(prompt, timeout=240)
        return result if result else "（輪1生成失敗）"
    
    def _round2_verify_lawsuit(self, generated_lawsuit: str, accident_facts: str, compensation_text: str, parties: dict) -> str:
        """輪2：檢查驗證 Prompt"""
        
        verification_prompt = f"""你現在是一名檢查員，負責驗證上一步生成的起訴狀是否符合要求。請逐條回答，若不合格指出問題並給出修正方案。

📄 待檢查的起訴狀：
{generated_lawsuit}

📄 原始資料：
當事人：{parties}
事故事實：{accident_facts}
損害內容：{compensation_text}

檢查清單：

1. 事實段是否僅描述事故經過，未帶入傷勢或金額？

2. 法律依據段是否引用所有應用條文？（依上述判斷條件檢查184/191-2/193/195/185/188/187/190）

3. 185 / 187 / 188 條是否遵循優先規則？
   - 若有未成年被告+法定代理人 → 必須引用187，且不得引用185
   - 若有僱用關係 → 必須引用188，且不得引用185
   - 僅當無187/188且有多被告 → 引用185
   - 不得同時引用185與187 / 188，並必須說明選擇理由

4. 條文引用格式是否正確（文字先引用，再列條號）？

5. 損害項目段是否涵蓋輸入中所有金額？模糊描述（如「非財產損害」「仰賴他人照護」）是否正確歸類為精神慰撫金或看護費？

6. 損害項目段是否每項金額都有合理請求理由？總額是否與各項加總一致？

7. 文書是否包含四段且格式完整？

請逐項檢查並回答「通過」或「不通過」，若不通過請說明問題。
若任何一項為否，請重寫整篇起訴狀並修正，直到全部通過。

檢查結果："""

        verification_result = self.call_llm(verification_prompt, timeout=240)
        
        # 如果檢查發現問題，進行修正
        if "不通過" in verification_result or "重寫" in verification_result:
            print("⚠️ 檢查發現問題，進行修正...")
            correction_prompt = f"""基於以下檢查結果，請重寫完整的起訴狀：

檢查結果：
{verification_result}

原始起訴狀：
{generated_lawsuit}

請根據檢查意見完全重寫起訴狀，確保：
1. 修正所有指出的問題
2. 保持四段式結構完整
3. 法條引用完全正確
4. 金額計算準確無誤

重寫的起訴狀："""
            
            corrected_result = self.call_llm(correction_prompt, timeout=240)
            return corrected_result if corrected_result else generated_lawsuit
        
        return generated_lawsuit
    
    def _analyze_case_for_laws(self, parties: dict, accident_facts: str, compensation_text: str) -> str:
        """分析案件以確定適用法條"""
        analysis = []
        
        # 分析當事人狀況
        defendant_info = parties.get("被告", "")
        plaintiff_info = parties.get("原告", "")
        
        # 檢查未成年+法定代理人
        if "未成年" in defendant_info and "法定代理人" in defendant_info:
            analysis.append("✓ 被告含未成年人及法定代理人 → 應引用民法187條1項，不得引用185條")
        
        # 檢查僱用關係
        if "受僱" in defendant_info or "雇主" in defendant_info or "公司" in defendant_info:
            analysis.append("✓ 存在僱用關係 → 應引用民法188條1項，不得引用185條")
        
        # 檢查多被告
        defendant_count = parties.get("被告數量", 1)
        if defendant_count > 1:
            analysis.append("✓ 多名被告 → 可能適用民法185條1項（需確認無187/188條適用）")
        
        # 檢查損害類型
        if "精神" in compensation_text or "慰撫金" in compensation_text:
            analysis.append("✓ 有精神慰撫金請求 → 應引用民法195條1項前段")
        
        if "薪資" in compensation_text or "工作" in compensation_text or "看護" in compensation_text:
            analysis.append("✓ 有工作損失或看護費用 → 應引用民法193條1項")
        
        if "車禍" in accident_facts or "交通" in accident_facts:
            analysis.append("✓ 交通事故案件 → 應引用民法191條之2、184條1項前段")
        
        return "\n".join(analysis) if analysis else "無特殊法條適用狀況"
    
    def generate_complete_dual_round_cot(self, accident_facts: str, compensation_text: str, parties: dict, similar_cases: list = None, case_ids: list = None) -> dict:
        """完整的雙輪CoT起訴狀生成（包含檢索、統計、生成、驗證）"""
        
        # 生成事實段落
        print("📝 生成事實段落...")
        facts = self.generate_standard_facts(accident_facts, similar_cases)
        print("✅ 事實段落生成完成")
        
        # 統計相似案例法條並生成法律依據
        print("⚖️ 生成法律依據...")
        if similar_cases and case_ids:
            try:
                print("📊 分析相似案例使用的法條...")
                similar_laws_stats = get_similar_cases_laws_stats(case_ids)
                if similar_laws_stats:
                    print("📋 相似案例常用法條統計:")
                    for law_name, count in similar_laws_stats[:5]:
                        print(f"   • {law_name}: {count}次")
                    print()
            except Exception as e:
                print(f"⚠️ 法條統計分析失敗: {e}")
        
        laws = self.generate_standard_laws(
            accident_facts,
            "",  # injuries 留空
            parties,
            compensation_text
        )
        print("✅ 法律依據生成完成")
        
        # 生成損害賠償
        print("💰 生成損害賠償...")
        damages = self.generate_smart_compensation(
            "",  # injuries 留空
            compensation_text, 
            parties
        )
        print("✅ 損害賠償生成完成")
        
        # 使用雙輪CoT生成結論
        print("🧠 雙輪CoT結論生成...")
        conclusion = self._round1_generate_lawsuit(accident_facts, compensation_text, parties, similar_cases)
        verified_conclusion = self._round2_verify_lawsuit(conclusion, accident_facts, compensation_text, parties)
        print("✅ 雙輪CoT結論生成完成")
        
        print("\n✅ 所有生成步驟完成！")
        
        # 提取適用法條
        applicable_laws = determine_applicable_laws(
            accident_facts,
            "",  # injuries
            compensation_text,
            parties
        )
        
        # 組合完整起訴狀
        # Debug: Check variable types before concatenation
        if not isinstance(facts, str):
            print(f"❌ facts is not string, type: {type(facts)}, value: {facts}")
            facts = str(facts)
        if not isinstance(laws, str):
            print(f"❌ laws is not string, type: {type(laws)}, value: {laws}")
            laws = str(laws)
        if not isinstance(damages, str):
            print(f"❌ damages is not string, type: {type(damages)}, value: {damages}")
            damages = str(damages)
        if not isinstance(verified_conclusion, str):
            print(f"❌ verified_conclusion is not string, type: {type(verified_conclusion)}, value: {verified_conclusion}")
            verified_conclusion = str(verified_conclusion)
            
        full_lawsuit = f"""{facts}

{laws}

{damages}

{verified_conclusion}"""
        
        return {
            "lawsuit": full_lawsuit,
            "applicable_laws": applicable_laws,
            "facts": facts,
            "laws": laws,
            "damages": damages,
            "conclusion": verified_conclusion
        }
    
    def _build_structured_cot_prompt(self, accident_facts: str, compensation_text: str, parties: dict, analysis_result: dict) -> str:
        """構建基於結構化分析的CoT提示詞"""
        
        plaintiff = parties.get("原告", "原告")
        defendant = parties.get("被告", "被告")
        
        # 提取結構化分析結果
        structured_items = analysis_result.get('structured_items', [])
        calculation = analysis_result.get('calculation', {})
        validation = analysis_result.get('validation', {})
        
        # 構建項目摘要
        items_summary = "\n📋 已識別的損害項目："
        for item in structured_items:
            items_summary += f"\n• {item['item_title']}: {item['formatted_amount']}"
        
        items_summary += f"\n\n💰 計算分析："
        items_summary += f"\n• 正確總計: {calculation.get('total', 0):,}元"
        
        if validation.get('claimed_total'):
            items_summary += f"\n• 原起訴狀聲稱: {validation['claimed_total']:,}元"
            if validation.get('difference', 0) != 0:
                if validation['difference'] < 0:
                    items_summary += f"\n• ❌ 原起訴狀少算了: {abs(validation['difference']):,}元"
                else:
                    items_summary += f"\n• ❌ 原起訴狀多算了: {validation['difference']:,}元"
                items_summary += f"\n• ✅ 請使用正確金額: {calculation.get('total', 0):,}元"
        
        prompt = f"""你是台灣資深律師，請運用Chain of Thought推理方式，根據結構化分析結果生成專業的起訴狀結論段落。

🎯 重要指示：
1. 必須使用正確的金額計算，避免重複計算說明文字中的金額
2. 使用逐步推理方式分析損害項目
3. 結論必須包含完整的項目明細
4. 總金額必須準確無誤
5. 採用標準的法律文書格式

👥 當事人資訊：
原告：{plaintiff}
被告：{defendant}

📄 案件事實：
{accident_facts}

📄 損害賠償原始內容：
{compensation_text}

{items_summary}

🧠 請使用Chain of Thought方式分析：

步驟1: 分析案件性質和當事人責任
步驟2: 檢視各項損害的合理性和法律依據
步驟3: 驗證金額計算的準確性
步驟4: 綜合分析並形成結論

🏛️ 最後請生成專業的結論段落，包括：
1. 損害項目明細列表
2. 正確的總金額計算  
3. 標準的結論格式
4. 利息計算條款

格式要求：
- 開頭：「綜上所陳」
- 中間：列舉各項損害明細
- 結尾：總計金額和利息請求
- ⚠️ 重要：避免重複說明同一項目，每項損害只說明一次
- ⚠️ 重要：不要重複列舉金額，使用結構化分析的正確總額

重要：請確保金額計算絕對正確，使用結構化分析的正確總額！"""

        return prompt
    
    def _build_traditional_cot_prompt(self, accident_facts: str, compensation_text: str, parties: dict) -> str:
        """構建傳統CoT提示詞"""
        
        plaintiff = parties.get("原告", "原告")
        defendant = parties.get("被告", "被告")
        
        prompt = f"""你是台灣資深律師，請運用Chain of Thought推理方式生成專業的起訴狀結論段落。

👥 當事人資訊：
原告：{plaintiff}
被告：{defendant}

📄 案件事實：
{accident_facts}

📄 損害賠償內容：
{compensation_text}

🧠 請使用Chain of Thought方式分析：

步驟1: 分析案件性質和當事人責任
步驟2: 檢視各項損害的合理性和法律依據  
步驟3: 計算總損害金額
步驟4: 綜合分析並形成結論

🏛️ 最後請生成專業的結論段落，格式要求：
- 開頭：「綜上所陳」
- 中間：列舉各項損害明細
- 結尾：總計金額和利息請求
- ⚠️ 重要：避免重複說明同一項目，每項損害只說明一次
- ⚠️ 重要：不要重複列舉相同的金額和項目"""

        return prompt
    
    def _post_process_structured_conclusion(self, conclusion: str, analysis_result: dict) -> str:
        """後處理結構化結論"""
        
        # 添加處理資訊
        processing_info = f"\n\n💡 處理資訊：\n"
        processing_info += f"處理方法：結構化分析\n"
        
        calculation = analysis_result.get('calculation', {})
        validation = analysis_result.get('validation', {})
        
        processing_info += f"正確總額：{calculation.get('total', 0):,}元\n"
        if validation.get('claimed_total'):
            processing_info += f"原聲稱額：{validation['claimed_total']:,}元\n"
            if validation.get('difference', 0) != 0:
                processing_info += f"差額修正：{abs(validation['difference']):,}元\n"
        
        return conclusion + processing_info

    def _generate_complex_compensation(self, comp_facts: str, parties: dict) -> str:
        """處理複雜損害項目文本的分步方法"""
        print("🔧 使用分步處理複雜損害文本...")
        
        # 步驟1：預處理中文數字
        preprocessed_facts = self._preprocess_chinese_numbers(comp_facts)
        
        # 直接格式化為標準損害項目
        format_prompt = f"""請將以下損害賠償內容重新整理為標準的法律文書格式：

【當事人】
原告：{parties.get('原告', '未提及')}（共{parties.get('原告數量', 1)}名）

【損害描述】
{preprocessed_facts}

【標準格式要求】
（一）原告[姓名]之損害：
1. [項目名稱]：[金額]元
   說明：原告[姓名]因本次車禍[損害性質]
2. [項目名稱]：[金額]元
   說明：原告[姓名]因本次車禍[損害性質]

（二）原告[姓名]之損害：
1. [項目名稱]：[金額]元
   說明：原告[姓名]因本次車禍[損害性質]

【重要要求】
- 直接整理損害項目，不要顯示分析或計算過程
- 共同費用要平均分攤給相關原告
- 所有金額使用千分位逗號格式
- 每項損害都要有具體說明
- 確保格式整齊統一
- ⚠️ **姓名完整性**：必須完整保留原告的姓名，不可遺漏任何字（例如：「羅靖崴」絕對不能寫成「羅崴」，「陳皆宏」不能寫成「陳宏」）

【細項處理規則】⭐ 新增
- ✅ **保留細項結構**：如果律師輸入中某個損害項目有細項編號（如⑴⑵⑶或1.2.3.），請保留這個結構
- ✅ **範例**：
  輸入：「2.就醫交通費用：⑴計程車10,890元：往返醫院復健 ⑵救護車1,900元：出院時使用」
  輸出：
  ```
  （二）就醫交通費用：12,790元
  1. 計程車車資：10,890元
     往返醫院復健
  2. 救護車費用：1,900元
     出院時使用
  ```
- ✅ **合併金額**：大項目的金額是所有細項的總和
- ✅ **完整引用原文**：每個細項的說明要完整引用律師原文，不要改寫、不要簡化

請直接輸出標準格式的損害項目："""

        result = self.call_llm(format_prompt, timeout=120)

        # 修正可能被LLM截斷的原告姓名
        if result:
            result = self._fix_plaintiff_names(result, parties)

        return result

    def _verify_calculation(self, result_text: str) -> dict:
        """從結果中提取並驗證計算準確性"""
        import re
        
        verification = {
            "correct": True,
            "errors": [],
            "corrected_total": None
        }
        
        # 提取所有金額數字
        amounts = re.findall(r'(\d{1,3}(?:,\d{3})*)', result_text)
        amounts = [int(amt.replace(',', '')) for amt in amounts if amt]
        
        if len(amounts) >= 10:  # 如果有足夠的金額進行驗證
            # 嘗試找到兩個小計和總計
            try:
                # 假設最後三個大數字是：小計1、小計2、總計
                if len(amounts) >= 3:
                    subtotal1 = amounts[-3]
                    subtotal2 = amounts[-2] 
                    reported_total = amounts[-1]
                    
                    # 驗證總計
                    actual_total = subtotal1 + subtotal2
                    if actual_total != reported_total:
                        verification["correct"] = False
                        verification["errors"].append(f"總計錯誤：{subtotal1} + {subtotal2} = {actual_total}，但報告為{reported_total}")
                        verification["corrected_total"] = actual_total
                        
            except Exception as e:
                verification["errors"].append(f"驗證過程出錯：{e}")
        
        return verification

    def _fix_incomplete_sentences(self, text: str) -> str:
        """修復損害項目中被截斷的句子，補上標準證據說明"""
        import re

        # 定義常見的截斷模式和對應的標準補充句
        incomplete_patterns = [
            # 模式1: "...據。" → 補上完整的證據說明（必須是被截斷的）
            (r'([，、])\s*業據提[^\n。]*?據。(?![^\n]*為證)', r'\1業據提出相關診斷證明書、收據等資料為證。'),
            (r'業據[^\n。]*?據。(?![^\n]*為證)', r'業據提出相關證明文件為證。'),
            (r'([，、])\s*並提[^\n。]*?據。(?![^\n]*為證)', r'\1並提出相關診斷證明書、收據等資料為證。'),
            (r'並提[^\n。]*?據。(?![^\n]*為證)', r'並提出相關證明文件為證。'),
            (r'([，、])\s*並可提[^\n。]*?據。(?![^\n]*為證)', r'\1並可提出相關估價單、收據等資料為證。'),
            (r'並可提[^\n。]*?據。(?![^\n]*為證)', r'並可提出相關證明文件為證。'),

            # 模式2: "...證。" 但只在明顯被截斷的情況（如「作為證。」）
            (r'作為證(?![明據])[。，]', r'作為證據。'),

            # 模式3: 只有"據。"沒有其他內容
            (r'(?<=[，、。])\s*據。', r'有相關證明文件可資佐證。'),

            # 模式4: 金額後面直接跟"據。"
            (r'(\d+[,\d]*元)[部分]*\s*[。，]\s*據。', r'\1，有相關證明文件可資佐證。'),
        ]

        result = text
        fixes_applied = []

        for pattern, replacement in incomplete_patterns:
            matches = list(re.finditer(pattern, result))
            if matches:
                for match in matches:
                    fixes_applied.append(f"修復: '{match.group()[:30]}...' → '{replacement[:30]}...'")
                result = re.sub(pattern, replacement, result)

        # Debug: 顯示修復了哪些內容
        if fixes_applied:
            print("\n🔧 自動修復不完整句子:")
            for fix in fixes_applied[:5]:  # 只顯示前5個
                print(f"  • {fix}")
            if len(fixes_applied) > 5:
                print(f"  ... 以及其他 {len(fixes_applied) - 5} 處修復")

        return result

    def _split_compensation_facts_into_items(self, text: str) -> str:
        """將律師原文按照損害項目拆分成多段,方便 LLM 提取"""
        import re

        # 🎯 優先處理：如果是列表格式（1. 2. 3.），先按列表項拆分
        # 檢測模式：「1. XXX元」、「2. XXX元」等
        list_pattern = r'^\d+\.\s+.+?\d+[,\d]*元'
        lines = text.split('\n')
        has_numbered_list = sum(1 for line in lines if re.match(list_pattern, line.strip())) >= 3

        if has_numbered_list:
            # 這是列表格式，按照數字編號拆分
            paragraphs = []
            current_item = []
            in_list = False

            for line in lines:
                line_stripped = line.strip()
                # 如果是新的編號項（如「1. 」、「2. 」）
                if re.match(r'^\d+\.\s+', line_stripped):
                    # 保存前一個項目
                    if current_item:
                        paragraphs.append('\n'.join(current_item))
                    # 開始新項目
                    current_item = [line_stripped]
                    in_list = True
                elif line_stripped and in_list:
                    # 如果這行也是編號項的延續（包含括號說明等）
                    # 或者這行以「原告」開頭（可能是列表後的說明，應該停止）
                    if line_stripped.startswith('原告'):
                        # 結束列表，開始其他內容
                        in_list = False
                        if current_item:
                            paragraphs.append('\n'.join(current_item))
                            current_item = []
                    elif current_item:
                        # 繼續當前項目（括號說明等）
                        current_item.append(line_stripped)

            # 保存最後一個項目
            if current_item:
                paragraphs.append('\n'.join(current_item))

            if paragraphs:
                print(f"🔍 偵測到列表格式，已拆分為 {len(paragraphs)} 個項目")
                return '\n\n'.join(paragraphs)

        # 先統一段落分隔符:將單個\n轉換為\n\n(如果該換行前後都有文字)
        # 這樣可以保留原有的段落結構
        text = re.sub(r'([。！？])\s*\n\s*([原又系最])', r'\1\n\n\2', text)

        # 如果原文本來就有分段(包含連續的換行),檢查是否需要進一步拆分慰撫金
        if text.count('\n\n') >= 2:
            # 清理多餘的空白
            paragraphs = [p.strip() for p in text.split('\n\n') if p.strip()]

            # 檢查每個段落是否包含慰撫金且需要進一步拆分
            refined_paragraphs = []
            for para in paragraphs:
                # 如果段落包含多個金額且有慰撫金，需要拆分
                amounts = re.findall(r'\d+[,\d]*元', para)
                has_solatium = '慰撫金' in para

                if len(amounts) > 2 and has_solatium:
                    # 這個段落需要進一步拆分（在慰撫金前面拆分）
                    # 找到慰撫金前的最後一個「元。」或「元，」
                    solatium_pos = para.find('慰撫金')
                    if solatium_pos > 0:
                        before_solatium = para[:solatium_pos]
                        # 找最後一個「元。」
                        last_period = before_solatium.rfind('元。')
                        if last_period != -1:
                            split_point = last_period + 2
                            part1 = para[:split_point].strip()
                            part2 = para[split_point:].strip()
                            refined_paragraphs.append(part1)
                            refined_paragraphs.append(part2)
                        else:
                            refined_paragraphs.append(para)
                    else:
                        refined_paragraphs.append(para)
                else:
                    refined_paragraphs.append(para)

            return '\n\n'.join(refined_paragraphs)

        # 使用更精確的拆分策略:
        # 1. 按照「。而原告」、「。並且原告」、「。原告再」、「。最後原告」等模式拆分
        # 2. 按照「元。」後面跟著「而」、「並且」、「原告」等詞拆分

        # 定義拆分點的模式（修正版：不要砍掉「原告」等詞）
        split_patterns = [
            # 🎯 優先處理慰撫金（最重要！幾乎每篇都有）
            # 模式：「元。末查...爰請求精神慰撫金」→ 在「元。」後拆分，但保留所有慰撫金背景說明
            (r'(元[。，])\s*(末查|末按|查原告)', r'\1\n\n\2'),  # "元。末查..." → "元。\n\n末查..."

            # 原有的拆分模式
            (r'(元[。，])\s*(另查|另按|又原告)', r'\1\n\n\2'),  # "元。另查..." → "元。\n\n另查..."
            (r'(元[。，])\s*((?:而|並且|且|另|再者|又|復|另外|再))', r'\1\n\n\2'),  # "元。而..." → "元。\n\n而..."
            (r'(元[。，])\s*(原告)', r'\1\n\n\2'),  # "元。原告..." → "元。\n\n原告..."
            (r'([。！])\s*((?:而|並且|且|另|再者|又|復|另外|再)原告)', r'\1\n\n\2'),  # "。而原告..." → "。\n\n而原告..."
            (r'([。！])\s*(最後原告)', r'\1\n\n\2'),  # "。最後原告..." → "。\n\n最後原告..."
        ]

        result_text = text
        for pattern, replacement in split_patterns:
            # 在匹配的位置插入段落標記，但保留被匹配到的詞
            result_text = re.sub(pattern, replacement, result_text)

        # 如果成功拆分(包含段落分隔),返回拆分後的文本
        if '\n\n' in result_text:
            # 清理多餘的空白行
            paragraphs = [p.strip() for p in result_text.split('\n\n') if p.strip()]
            return '\n\n'.join(paragraphs)
        else:
            # 如果上述模式都沒匹配到,嘗試按照「XX元」後面跟著「XX費用」的模式拆分
            # 例如: "...190,948元，並且有醫療費用收據...看護費用..." → "元" 和 "看護費用" 之間拆分
            pattern = r'(\d+[,\d]*元[，。、；].*?)(?=\d+[,\d]*元|$)'
            matches = re.findall(pattern, result_text, re.DOTALL)
            if len(matches) > 1:
                return '\n\n'.join([m.strip() for m in matches if m.strip()])

        # 拆分失敗,返回原文
        return text

    def _generate_llm_based_compensation(self, comp_facts: str, parties: dict) -> str:
        """使用LLM完全處理損害項目生成"""

        # 先預處理中文數字
        preprocessed_facts = self._preprocess_chinese_numbers(comp_facts)

        # 將一整段文字拆分成多個段落,方便 LLM 提取
        preprocessed_facts = self._split_compensation_facts_into_items(preprocessed_facts)

        # Debug: 輸出拆分後的段落
        print("=" * 60)
        print("📋 拆分後的損害描述段落:")
        print("=" * 60)
        for i, para in enumerate(preprocessed_facts.split('\n\n'), 1):
            print(f"\n段落 {i}:")
            print(para[:200] + ('...' if len(para) > 200 else ''))
        print("=" * 60)
        
        # 檢查是否為單一原告情況（無論被告數量）
        plaintiff_count = parties.get('原告數量', 1)
        defendant_count = parties.get('被告數量', 1)
        is_single_plaintiff = plaintiff_count == 1

        if is_single_plaintiff:
            # 單一原告時，使用中文編號格式（無論被告數量）
            defendant_info = f"{parties.get('被告', '被告')}（共{defendant_count}名）" if defendant_count > 1 else f"{parties.get('被告', '被告')}（單一被告）"

            # 統計原始描述中的段落數量(每個包含金額的段落)
            import re
            amount_paragraphs = [p for p in preprocessed_facts.split('\n\n') if re.search(r'\d+[,\d]*元', p)]
            expected_item_count = len(amount_paragraphs)

            # 生成段落提示清單
            paragraph_list = []
            for i, para in enumerate(amount_paragraphs, 1):
                # 提取第一個出現的金額作為提示
                first_amount_match = re.search(r'(\d+[,\d]*)元', para)
                first_amount = first_amount_match.group(1) if first_amount_match else "?"
                # 截取前50字作為提示
                preview = para[:50].replace('\n', ' ') + ('...' if len(para) > 50 else '')
                paragraph_list.append(f"{i}. 金額包含{first_amount}元 - {preview}")

            paragraph_hints = '\n'.join(paragraph_list)

            prompt = f"""【任務】將律師提供的損害描述重新格式化成起訴狀格式

🔥 **核心要求**：
1. 原始描述有 {expected_item_count} 個段落包含金額 → 你必須生成 {expected_item_count} 個損害項目
2. 每個段落的文字要「逐字逐句照搬」,不要改寫、不要簡化、不要刪減
3. 只調整編號格式(改成「（一）、（二）...」),其他內容**一字不改**

⚠️ **針對簡短列表項的特殊處理**：
如果律師只提供簡短的項目名稱和金額（如「1. 醫療費用28,170元」），你應該：
- ✅ 保留項目名稱和金額
- ✅ 加上標準的法律用語補充（如「原告因本次事故受有...」）
- ✅ 如果有括號說明（如「(原告主張需要休養和復健治療3個月)」），必須完整保留

例如：
原文：「5. 因無法工作造成的損失90,000元(原告主張需要休養和復健治療3個月)」
應輸出：
（五）工作損失：90,000元
原告因本次事故無法工作，受有損失90,000元（原告主張需要休養和復健治療3個月）。

⚠️ **核心原則：保留所有資訊和細節，但可以調整格式讓文字更易讀**

假設律師原文是：
"原告主張根據診斷證明書記載原告所受傷勢，因傷及足部而有礙行動，因此有搭乘計程車之必要性，參考網路查詢自原告住家至新光醫院、太順中醫診所、蔡嘉哲骨科診所、六善堂中醫診所之單程計程車車資各為455元、85元、165元、185元後，總計支出往返就醫交通費25,885元。"

✅ **正確**(保留所有細節，格式調整讓更易讀)：
（X）交通費用：25,885元
原告主張根據診斷證明書記載原告所受傷勢，因傷及足部而有礙行動，因此有搭乘計程車之必要性，參考網路查詢自原告住家至新光醫院、太順中醫診所、蔡嘉哲骨科診所、六善堂中醫診所之單程計程車車資各為455元、85元、165元、185元後，總計支出往返就醫交通費25,885元。

❌ **錯誤**(刪減了重要細節)：
（X）交通費用：25,885元
原告因本次事故受有交通費25,885元。← 刪掉了醫院名稱和各個車資!這是錯的!

【原始損害描述】
{preprocessed_facts}

【段落清單】你必須為以下每個段落生成一個損害項目：
{paragraph_hints}

【輸出格式】
（一）[項目名稱]：[金額]
[引用對應段落的原文,保留**所有重要資訊**：醫院/診所名稱、具體金額明細、計算式(如26,000×6個月)、證據說明(如「有估價單可作為證據」、「並提出收據為證」)等。可以適度調整格式讓文字更易讀,但絕對不能刪減任何資訊!]

（二）[項目名稱]：[金額]
[引用對應段落的原文,保留**所有重要資訊**：醫院/診所名稱、具體金額明細、計算式、證據說明等。可以適度調整格式,但絕對不能刪減任何資訊!]

...依此類推,直到第 {expected_item_count} 項

🚨 **絕對禁止**：
- ❌ 不要遺漏任何段落
- ❌ 不要刪減資訊(醫院名稱、金額明細、計算式、證據說明等必須保留)
- ❌ 不要截斷句子(要把句子寫完整,如「有估價單可作為證據」不能只寫到「據」就停了)
- ❌ 不要自己編造內容

✅ **允許的調整**：
- ✅ 可以適度斷句讓文字更易讀
- ✅ 可以調整段落結構
- ✅ 但所有具體資訊(名稱、數字、計算式、證據)必須完整保留

現在開始輸出,記住:**把原文從頭到尾完整複製過來**!"""
        else:
            # 多原告或多被告時，使用完整格式，但每位原告內部使用中文編號

            # 統計原始描述中的段落數量(每個包含金額的段落)
            import re
            amount_paragraphs = [p for p in preprocessed_facts.split('\n\n') if re.search(r'\d+[,\d]*元', p)]
            expected_item_count = len(amount_paragraphs)

            prompt = f"""你是台灣律師，請根據車禍案件的損害賠償內容，分析並重新整理成標準的起訴狀損害項目格式：

【當事人資訊】
原告：{parties.get('原告', '未提及')}（共{parties.get('原告數量', 1)}名）
被告：{parties.get('被告', '未提及')}（共{parties.get('被告數量', 1)}名）

【原始損害描述】
{preprocessed_facts}

⚠️ **統計資訊**：原始描述中包含金額的段落共有 {expected_item_count} 個，所有原告的損害項目總數也必須正好是 {expected_item_count} 個

🚨 **重要警告**：
- ❌ **絕對禁止從下方的範例中抄金額**！範例中的所有金額都是示意用的，不是本案的實際金額！
- ✅ **所有金額和描述都必須從上方的「原始損害描述」中提取**
- ✅ 範例只是用來說明格式，不是讓你抄內容的

🚨 **原告姓名規則（最重要！）**：
- ✅ **只能使用「當事人資訊」中列出的原告姓名**，本案原告為：{parties.get('原告', '原告')}
- ❌ **絕對禁止使用匿名格式**：不可寫「賴○○」、「陳○○」、「林○○」等有○○的寫法，必須寫出完整真實姓名
- ❌ **絕對禁止憑空生成原告**：輸出的原告人數必須恰好等於 {parties.get('原告數量', 1)} 人，不可多也不可少
- ❌ **絕對禁止使用範例中的假名**：不可用「甲」「乙」「XXX」「某某某」等，必須用真實姓名

🚨 **重要識別規則**：
- ✅ **只處理原告的損害**：只能為明確標示為「原告XXX」的人生成損害項目
- ❌ **絕對排除訴外人**：標示為「訴外人XXX」的人不是原告，不要為他們生成任何損害項目
- ❌ **排除乘客身份**：如果某人只被描述為「乘客」、「搭載」、「車上乘客」而沒有「原告」二字，不要為他們生成損害項目
- ✅ **正確範例**：
  * 「原告陳皆宏駕駛...另一原告王惠華」→ 為陳皆宏和王惠華各自生成損害項目
- ❌ **錯誤範例**：
  * 「車上搭載乘客訴外人王惠滿」→ 不要為王惠滿生成損害項目（即使原始描述中有王惠滿的醫療費用等資料）
  * 如果原始描述提到「原告王惠滿」，但當事人資訊中沒有王惠滿，也不要生成

【分析要求】
請仔細分析上述內容，從中提取出：
1. **完整性要求**：必須提取原始描述中的**所有**損害項目，不可遺漏任何一項
2. 每位原告的具體損害項目類型和確切金額
3. 每項損害的事實根據和法律理由
4. **重要**：只能使用原始描述中已提及的事實，絕對不可以自行添加或編造任何內容
5. **重要**：如果原始描述中有分成多個小項的大項目（如「醫療費用部分合計8萬3,016元」下面有多個細項），請保持整合為一個大項，在說明中詳列細項明細
6. **重要**：原始描述中的每一個編號項目（如：1. XXX、2. XXX、3. XXX...）都必須在輸出中出現，絕不可遺漏
7. **重要**：慰撫金、減少勞動能力損失等大額項目特別容易被遺漏，請務必檢查並包含
8. **重要**：只為「當事人資訊」中列出的原告生成損害項目，絕對不要為「訴外人」生成損害項目

⚠️ **逐段提取原則**（⭐ 非常重要）：
原始描述中的每一段文字如果包含金額，就必須對應一個獨立的損害項目！

**步驟1**: 先從「原始損害描述」中逐段找出所有包含金額的段落（向上滾動仔細檢查）
**步驟2**: 為每個段落確定損害項目名稱和金額
**步驟3**: 按順序生成損害項目

🔍 **自我檢查**：
- 在生成前，先數一數「原始損害描述」中有幾個段落包含金額？
- 你的輸出項目數量必須等於這個數字
- 如果不相等，說明你遺漏了某些段落，請重新檢查「原始損害描述」

範例:
- ✅ 如果看到「支出相關就醫費用共130,698元」→ 生成「醫療費用：130,698元」
- ✅ 如果看到「預估仍有將來醫療費用共150,000元」→ 生成「預估未來醫療費用：150,000元」
- ✅ 如果看到「支出往返就醫交通費25,885元」→ 生成「就醫交通費用：25,885元」
- ✅ 如果看到「購買...醫療用品，支出45,749元」→ 生成「醫療用品費用：45,749元」
- ✅ 如果看到「購買...營養品支出18,000元」→ 生成「營養品費用：18,000元」
- ✅ 如果看到「不能工作...損失金額為156,000元」→ 生成「不能工作損失：156,000元」
- ✅ 如果看到「支付維修費用13,750元」→ 生成「車輛維修費用：13,750元」
- ✅ 如果看到「勞動能力減損之損害200,000元」→ 生成「勞動能力減損：200,000元」
- ✅ 如果看到「精神慰撫金600,000元」→ 生成「慰撫金：600,000元」

⚠️ 注意：「不能工作損失」和「勞動能力減損」是**兩個不同的項目**，不要合併！
⚠️ 注意：「車輛維修費」和「車輛貶值」是**兩個不同的項目**，不要合併！
⚠️ 注意：每個包含金額的段落都要變成一個損害項目，一個都不能漏！

⚠️ **完整性檢查清單**（輸出前必須確認）：
□ 是否原始描述中的每個段落（包含金額）都已轉換為一個損害項目？
□ 是否包含了「減少勞動能力損失」或「不能工作損失」（如果原始描述中有此項）？
□ 是否包含了「勞動能力減損」（如果原始描述中有此項，且它與「不能工作損失」是分開的）？
□ 是否包含了「車輛維修費用」（如果原始描述中有此項）？
□ 是否包含了「慰撫金」或「精神慰撫金」（如果原始描述中有此項）？
□ 每位原告的輸出項目數量是否 = 原始描述中該原告包含金額的段落數量？（不多不少）
□ 是否有任何項目標註為「未提供」？（如果有，請刪除該項目）

【標準輸出格式範例】（完整引用原文）

假設律師輸入是：
「（一）原告甲之損害：
1. 醫療費用：50,000元
原告甲因本次事故受傷就醫，支出醫療費用50,000元。

2. 慰撫金：300,000元
原告甲因本次車禍，故請求精神慰撫金300,000元。

（二）原告乙之損害：
1. 醫療費用：30,000元
原告乙因本次事故受傷就醫，支出醫療費用30,000元。」

應輸出：
（一）原告甲之損害：
1. 醫療費用：50,000元
原告甲因本次事故受傷就醫，支出醫療費用50,000元。

2. 慰撫金：300,000元
原告甲因本次車禍，故請求精神慰撫金300,000元。

（二）原告乙之損害：
1. 醫療費用：30,000元
原告乙因本次事故受傷就醫，支出醫療費用30,000元。

⚠️ 關鍵：每個項目的說明都要完整引用律師原文，一字不改!

⚠️ **細項處理原則**（⭐ 重要）：
- 如果律師輸入中某個損害項目有細項編號（如⑴⑵⑶），請在說明中保留這個細項結構
- 範例：如果律師輸入是「2.就醫交通費用：⑴計程車10,890元 ⑵救護車1,900元」
  應輸出：
  ```
  2. 就醫交通費用：12,790元
  原告因本次事故就醫，支出交通費用如下：⑴計程車車資10,890元（需搭乘計程車往返醫院與住家就診）、⑵救護車費用1,900元（住院出院返家時支出），合計12,790元。
  ```
- 大項標題使用所有細項的總和
- 說明中列出每個細項的編號、名稱、金額和描述
- 每個細項的描述要從律師原文中提取，保持忠實

【關鍵要求】
- 每位原告先用（一）（二）等編號區分
- 每位原告內部的損害項目使用 1. 2. 3. 等數字編號
- 每項格式：數字編號. 項目名稱：總金額
  **下一行必須詳細說明該項損害的具體事實和理由**
- ⚠️ **重要**：每個損害項目（除了標題行外）都必須有詳細的說明段落
- ⚠️ **重要項目整合原則**：
  * 如果原始描述將某類費用（如醫療費用）分成多個細項，請整合成一個大項
  * 在該大項的說明中，詳細列出所有細項的明細和金額
  * 標題行使用總金額，說明段落中列出細項分別的金額
- ⚠️ **慰撫金必須列出**：如果原始描述中有提到慰撫金（或精神慰撫金），必須作為獨立項目列出
- 理由說明必須盡可能引用原始描述中的具體事實、數字、細節
- 例如：如果原始描述提到「住院開刀醫療費用5萬3,122元」，就要完整引用這些細節
- 不可自行編造任何醫療診斷、傷勢描述或其他細節
- 理由要採用正式的法律文書語言
- 使用千分位逗號格式顯示金額

【嚴格禁止事項】
- ❌ **絕對不可生成原始描述中沒有的項目**：
  * 如果原始描述中沒有提到某個損害項目，就不要生成該項目
  * 不要生成「未提供資訊」、「本案未提供」等標註
  * 不要使用固定模板列出所有可能的項目類型
  * 只生成原始描述中**實際存在且有金額**的損害項目
- 絕對不可在輸出中包含「綜上所述」、「總計」、「合計」、「共計」等結論性文字
- 不要包含任何總金額計算或匯總說明
- 不要包含任何法定利息的說明
- 不要包含任何結論段落或總結文字
- 不要包含證據相關文字：「此有相關收據可證」、「有收據為證」、「有統一發票可證」、「可證」等
- 不要包含判決書用語：「經查」、「查明」、「經審理」等
- 只輸出純粹的損害項目條列，每項包含編號、名稱、金額、理由說明

⚠️ **最重要的格式要求**：
每個損害項目都必須包含兩個部分：
1. 第一行：數字編號. 項目名稱：金額
2. 第二行起：具體的事實說明和理由

🔥 **說明段落的寫法**（⭐⭐⭐ 極度重要）：
- ✅ **幾乎完整引用律師原文**：不要改寫、不要簡化、不要自己編造
- ✅ 保留原文的用詞、句式、細節
- ❌ 不要寫成「原告因傷就醫，支出醫療費用」這種簡化版
- ✅ 要寫成律師原文的樣子，例如：「原告因傷不良於行，上下班須搭乘計程車，支出交通費用18,000元」

❌ **絕對禁止**：只有標題行沒有說明的情況，例如：
1. 醫療費用：XXX元
2. 交通費用：XXX元  ← 這樣是錯誤的！沒有說明段落

⚠️ **檢查清單**：在輸出前，請確認：
□ 是否每位原告的每個項目都有說明段落？
□ 是否引用了原始描述中的具體細節？
□ 是否包含了慰撫金項目？

📋 **輸出前最後檢查 - 對照表**：

❌ 錯誤範例（缺少說明）：
（一）原告XXX之損害：
1. 醫療費用：XXX元
2. 看護費用：XXX元
3. 慰撫金：XXX元

✅ 正確範例（詳細引用原文）：

假設律師輸入是：「原告因傷不良於行，上下班須搭乘計程車，支出交通費用18,000元。」
應該生成：
（一）原告XXX之損害：
1. 交通費用：18,000元
原告因傷不良於行，上下班須搭乘計程車，支出交通費用18,000元。

假設律師輸入是：「原告因本次車禍造成勞動能力減損，請求減損賠償200,000元。」
應該生成：
2. 勞動能力減損：200,000元
原告因本次車禍造成勞動能力減損，請求減損賠償200,000元。

⚠️ **關鍵重點**：說明段落要「幾乎完整引用」律師原文的描述，不要改寫、不要簡化！

請嚴格按照上述格式和要求，基於原始描述的事實分析並輸出損害項目："""

        result = self.call_llm(prompt, timeout=120)

        # 清理結論性文字
        result = self._remove_conclusion_phrases(result)

        # 修復不完整的句子（如「據。」→「有相關證明文件可資佐證。」）
        result = self._fix_incomplete_sentences(result)

        # 修正 LLM 生成的匿名格式（如「賴○○」→「賴秀雲」）
        result = self._fix_anonymized_names(result, parties)

        # 檢查並補充缺少描述的項目
        result = self._ensure_all_items_have_description(result, preprocessed_facts)

        # 驗證並修正金額計算錯誤
        result = self._verify_and_fix_amount_calculations(result)

        # 格式化所有金額為千分位格式
        result = self._fix_mixed_amount_format(result)

        # 檢查結果是否包含預期格式
        if "（一）" in result and "原告" in result:
            # 清理結果，移除"三、損害項目："標題
            result = re.sub(r'^三、損害項目：\s*\n?', '', result)
            return result
        else:
            # Fallback：基本格式化
            return comp_facts

    def _verify_and_fix_amount_calculations(self, result: str) -> str:
        """驗證並修正損害項目中的金額計算錯誤"""
        import re

        lines = result.split('\n')
        output_lines = []
        i = 0

        while i < len(lines):
            line = lines[i]

            # 檢測包含細項明細的醫療費用行
            item_match = re.match(r'^(\d+)\.\s*(醫療費用|看護費用)：([\d,]+)元\s*$', line.strip())

            if item_match:
                item_num = item_match.group(1)
                item_name = item_match.group(2)
                claimed_amount = int(item_match.group(3).replace(',', ''))

                # 檢查下一行的描述中是否包含細項金額
                if i + 1 < len(lines):
                    description = lines[i + 1]

                    # 檢查是否包含計算式（如：2,400元×90天、2,200元×3個月）
                    has_calculation_formula = re.search(r'[\d,]+元\s*[×x]\s*[\d,]+\s*[天月]', description)

                    # 如果包含計算式，跳過驗證（因為是單價×數量的說明）
                    if has_calculation_formula:
                        output_lines.append(line)
                        i += 1
                        continue

                    # 檢查是否有明確的總計標記（如：「合計XXX元」、「計算式：...=XXX元」）
                    has_explicit_total = re.search(r'(合計|共計|共為|總計|計算式：.*?=)\s*([\d,]+)元', description)
                    if has_explicit_total:
                        # 提取明確標記的總額
                        explicit_total = int(has_explicit_total.group(2).replace(',', ''))
                        # 如果明確總額與標題金額一致，說明標題是正確的，不需修正
                        if explicit_total == claimed_amount:
                            output_lines.append(line)
                            i += 1
                            continue
                        # 如果不一致，使用明確標記的總額作為正確金額
                        elif explicit_total != claimed_amount:
                            print(f"⚠️  發現金額不一致：{item_name}")
                            print(f"   標題金額：{claimed_amount:,}元")
                            print(f"   描述中的總額：{explicit_total:,}元")
                            print(f"   差異：{abs(explicit_total - claimed_amount):,}元")
                            # 修正標題行使用描述中的總額
                            corrected_line = f"{item_num}. {item_name}：{explicit_total:,}元"
                            output_lines.append(corrected_line)
                            print(f"✅ 已修正為：{explicit_total:,}元")
                            i += 1
                            continue

                    # 提取描述中的所有金額
                    detail_amounts = re.findall(r'(\d+(?:,\d{3})*)元', description)

                    # 過濾：如果最後一個金額等於聲稱總額，說明描述末尾已經有總計，不應重複累加
                    if detail_amounts and int(detail_amounts[-1].replace(',', '')) == claimed_amount:
                        # 只加總除了最後一個金額外的金額
                        detail_amounts = detail_amounts[:-1]

                    if len(detail_amounts) > 1:  # 有多個細項金額
                        # 計算實際總額
                        actual_total = sum(int(amt.replace(',', '')) for amt in detail_amounts)

                        if actual_total != claimed_amount:
                            print(f"⚠️  發現金額計算錯誤：{item_name}")
                            print(f"   聲稱總額：{claimed_amount:,}元")
                            print(f"   實際總額：{actual_total:,}元（細項：{detail_amounts}）")
                            print(f"   差異：{abs(actual_total - claimed_amount):,}元")

                            # 修正標題行的金額
                            corrected_line = f"{item_num}. {item_name}：{actual_total:,}元"
                            output_lines.append(corrected_line)
                            print(f"✅ 已修正為：{actual_total:,}元")
                            i += 1
                            continue

            output_lines.append(line)
            i += 1

        return '\n'.join(output_lines)

    def _ensure_all_items_have_description(self, result: str, original_facts: str) -> str:
        """檢查並補充缺少描述的損害項目"""
        import re

        lines = result.split('\n')
        output_lines = []
        i = 0
        current_plaintiff = None  # 追蹤當前處理的原告

        while i < len(lines):
            line = lines[i]
            output_lines.append(line)

            # 檢測原告標題行（如：（一）原告陳慶華之損害：）
            plaintiff_match = re.match(r'^[（(][一二三四五六七八九十]+[）)].*?原告([^\s之：]+)之損害', line)
            if plaintiff_match:
                current_plaintiff = plaintiff_match.group(1)
                print(f"🔍 當前處理原告：{current_plaintiff}")

            # 檢測損害項目標題行（支持兩種格式）
            # 格式1: 1. 醫療費用：20,185元（數字編號）
            # 格式2: （一）醫療費用：2,370元（中文編號）
            item_match_num = re.match(r'^(\d+)\.\s*(.+?)：([\d,]+)元\s*$', line.strip())
            item_match_chinese = re.match(r'^[（(]([一二三四五六七八九十百]+)[）)]\s*(.+?)：([\d,]+)元\s*$', line.strip())

            item_match = item_match_num or item_match_chinese
            if item_match:
                if item_match_num:
                    item_num = item_match.group(1)
                    item_name = item_match.group(2)
                    amount = item_match.group(3)
                else:  # item_match_chinese
                    item_num = item_match.group(1)  # 中文編號
                    item_name = item_match.group(2)
                    amount = item_match.group(3)

                # 檢查下一行是否有描述
                has_description = False
                if i + 1 < len(lines):
                    next_line = lines[i + 1].strip()
                    # 如果下一行不是空行，也不是另一個項目標題，則認為有描述
                    # 需要檢查兩種格式：數字編號和中文編號
                    is_next_item_num = re.match(r'^(\d+)\.\s*(.+?)：([\d,]+)元', next_line)
                    is_next_item_chinese = re.match(r'^[（(][一二三四五六七八九十百]+[）)]\s*(.+?)：([\d,]+)元', next_line)
                    is_plaintiff_section = re.match(r'^[（(][一二三四五六七八九十]+[）)].*?原告', next_line)

                    if next_line and not is_next_item_num and not is_next_item_chinese and not is_plaintiff_section:
                        has_description = True

                # 如果沒有描述，嘗試從原始輸入中提取
                if not has_description:
                    print(f"⚠️  檢測到缺少描述的項目：{current_plaintiff or '原告'} - {item_name} {amount}元")

                    # 從原始輸入中尋找相關描述（傳入當前原告名稱）
                    description = self._extract_description_from_original(item_name, amount, original_facts, current_plaintiff)

                    if description:
                        output_lines.append(description)
                        print(f"✅ 已補充描述：{description[:50]}...")
                    else:
                        # 生成通用描述（傳入當前原告名稱）
                        generic_desc = self._generate_generic_description(item_name, amount, current_plaintiff)
                        output_lines.append(generic_desc)
                        print(f"ℹ️  使用通用描述：{generic_desc[:50]}...")

            i += 1

        return '\n'.join(output_lines)

    def _extract_description_from_original(self, item_name: str, amount: str, original_facts: str, plaintiff_name: str = None) -> str:
        """從原始輸入中提取項目描述"""
        import re

        # 清理金額格式（移除逗號）
        clean_amount = amount.replace(',', '')

        # 關鍵詞映射
        keyword_map = {
            '醫療費用': ['醫療費用', '就醫', '治療', '醫院', '診所', '馬偕', '榮總'],
            '復健費用': ['復健', '復健診所', '物理治療'],
            '診斷證明書費用': ['診斷證明書', '證明書'],
            '住院期間必要物品及洗頭費用': ['住院', '病房', '必要物品', '洗頭'],
            '看護費用': ['看護', '照護', '照顧', '生活無法自理', '專人'],
            '薪資損失': ['薪資', '收入', '工作', '無法工作', '不能工作', '休養', '請假', '扣薪'],
            '工作損失': ['薪資', '收入', '工作', '無法工作', '不能工作', '休養', '請假', '扣薪'],
            '不能工作之損失': ['薪資', '收入', '工作', '無法工作', '不能工作', '休養', '請假', '扣薪', '家務勞動'],
            '減少勞動能力損失': ['減少勞動能力', '勞動能力', '失能', '殘廢', '勞工保險失能', '失能等級', '霍夫曼', '流動性泥作師', '體力工作', '後遺症'],
            '交通費用': ['交通', '計程車', '往返', '車資', '回診', '復健'],
            '機車修復費用': ['機車', '修復', '修理', '受損'],
            '車輛修復費用': ['機車', '車輛', '修復', '修理', '受損', '毀損'],
            '車輛損壞': ['機車', '車輛', '修復', '修理', '受損', '毀損'],
            '增加生活必要費用': ['助行器', '拐杖', '輪椅', '生活必要', '醫療器材'],
            '精神慰撫金': ['慰撫金', '精神', '痛苦', '恐懼', '難以入眠', '止痛藥', '無法恢復正常生活'],
        }

        # 獲取關鍵詞
        keywords = keyword_map.get(item_name, [item_name])

        # 在原始文本中搜尋包含金額和關鍵詞的句子
        sentences = re.split(r'[。；\n]', original_facts)

        for sentence in sentences:
            # 跳過標題行格式（如：1. 醫療費用11,240元：）
            if re.match(r'^\d+\.\s*.+?：\s*$', sentence.strip()):
                continue
            if re.match(r'^\d+\.\s*.+?[：:]\s*[\d,]+元\s*[：:]?\s*$', sentence.strip()):
                continue

            # 檢查是否包含金額和關鍵詞
            has_amount = clean_amount in sentence.replace(',', '')
            has_keyword = any(kw in sentence for kw in keywords)

            # 如果指定了原告名稱，優先匹配包含該原告的句子
            if plaintiff_name:
                has_plaintiff = plaintiff_name in sentence or f"原告{plaintiff_name}" in sentence
                if has_amount and has_keyword and has_plaintiff:
                    # 清理句子
                    cleaned = sentence.strip()
                    if cleaned and not cleaned.startswith('原告主張'):
                        # 確保以"原告XXX"開頭
                        if not cleaned.startswith('原告'):
                            cleaned = f"原告{plaintiff_name}{cleaned}"

                        # 清理不應出現在起訴狀中的保險給付相關內容
                        cleaned = self._remove_insurance_references(cleaned)

                        return cleaned

        # 如果沒有找到包含特定原告的句子，回退到一般搜尋
        for sentence in sentences:
            # 跳過標題行格式
            if re.match(r'^\d+\.\s*.+?：\s*$', sentence.strip()):
                continue
            if re.match(r'^\d+\.\s*.+?[：:]\s*[\d,]+元\s*[：:]?\s*$', sentence.strip()):
                continue

            has_amount = clean_amount in sentence.replace(',', '')
            has_keyword = any(kw in sentence for kw in keywords)

            if has_amount and has_keyword:
                # 清理句子
                cleaned = sentence.strip()
                if cleaned and not cleaned.startswith('原告主張'):
                    # 確保以正確的原告名稱開頭
                    if not cleaned.startswith('原告'):
                        if plaintiff_name:
                            cleaned = f"原告{plaintiff_name}{cleaned}"
                        else:
                            cleaned = f"原告{cleaned}"
                    elif plaintiff_name and not cleaned.startswith(f"原告{plaintiff_name}"):
                        # 替換錯誤的原告名稱
                        cleaned = re.sub(r'^原告[^\s，。；、]+', f'原告{plaintiff_name}', cleaned)

                    # 清理不應出現在起訴狀中的保險給付相關內容
                    cleaned = self._remove_insurance_references(cleaned)

                    return cleaned

        return ""

    def _remove_insurance_references(self, text: str) -> str:
        """移除描述中不應出現在起訴狀中的保險給付相關內容"""
        import re

        # 定義需要移除的pattern（順序很重要，從最具體到最一般）
        insurance_patterns = [
            r'，?除強制責任險已給付部分外，?',
            r'，?除保險給付[^，。；]*?外，?',
            r'，?除.*?險.*?給付.*?外，?',
            r'，?扣除強制責任險[\d,，]*?元，?',
            r'，?扣除保險給付[\d,，]*?元，?',
            r'，?扣除.*?險[\d,，]*?元，?',
            r'，?已由保險給付[\d,，]*?元，?',
            r'，?保險已給付[\d,，]*?元，?',
            r'，?強制險已給付[\d,，]*?元，?',
        ]

        cleaned = text
        for pattern in insurance_patterns:
            cleaned = re.sub(pattern, '', cleaned)

        # 清理可能產生的多餘逗號
        cleaned = re.sub(r'，+', '，', cleaned)
        # 清理開頭和結尾的逗號
        cleaned = re.sub(r'^，|，$', '', cleaned)
        # 清理「原告，」這樣的格式，改為「原告」
        cleaned = re.sub(r'(原告[^，]*?)，(\s*[支於])', r'\1\2', cleaned)

        # 將「尚支出」改為「支出」（因為「尚」暗示了前面有扣除）
        cleaned = re.sub(r'尚支出', '支出', cleaned)

        return cleaned

    def _fix_anonymized_names(self, text: str, parties: dict) -> str:
        """修正 LLM 生成的匿名格式（如「賴○○」→「賴秀雲」）"""
        plaintiffs_str = parties.get('原告', '')
        if not plaintiffs_str or plaintiffs_str == '原告':
            return text

        # 解析真實原告姓名
        real_names = [n.strip() for n in plaintiffs_str.replace('、', ',').split(',') if n.strip() and n.strip() != '原告']
        if not real_names:
            return text

        # 找出文本中所有「X○○」或「XX○○」格式的匿名姓名
        anon_pattern = re.compile(r'([\u4e00-\u9fff]{1,2})(○+)')
        found_anon = list(dict.fromkeys(m.group(0) for m in anon_pattern.finditer(text)))

        for anon_name in found_anon:
            surname = anon_pattern.match(anon_name).group(1)
            # 在真實姓名中找姓氏相符的
            matched = [n for n in real_names if n.startswith(surname)]
            if len(matched) == 1:
                text = text.replace(anon_name, matched[0])
                print(f"🔧 修正匿名姓名：{anon_name} → {matched[0]}")
            elif len(matched) > 1:
                # 多個同姓氏，用名字長度推斷（○的數量對應名字字數）
                num_circles = anon_pattern.match(anon_name).group(2).count('○')
                surname_len = len(anon_pattern.match(anon_name).group(1))
                target_len = surname_len + num_circles
                best = [n for n in matched if len(n) == target_len]
                if len(best) == 1:
                    text = text.replace(anon_name, best[0])
                    print(f"🔧 修正匿名姓名：{anon_name} → {best[0]}")

        # 移除文本中出現但不在 parties 中的假原告段落標題行
        # 例如：「（二）原告陳○○之損害：」如果陳○○不是真實原告，整段刪除
        expected_count = parties.get('原告數量', 1)
        section_pattern = re.compile(
            r'[（(][一二三四五六七八九十]+[）)]\s*原告([^\s之：\n]{2,4})之損害[：:][^\n]*\n',
            re.MULTILINE
        )
        found_sections = section_pattern.findall(text)
        if found_sections and len(found_sections) > expected_count:
            # 有多餘的假原告段落，找出不在 real_names 中的
            fake_names = [n for n in found_sections if n not in real_names]
            if fake_names:
                print(f"⚠️ 偵測到假原告段落，嘗試移除：{fake_names}")

        return text

    def _fix_plaintiff_names(self, text: str, parties: dict) -> str:
        """修正文本中被LLM截斷的原告姓名"""
        import re

        # 獲取正確的原告姓名列表
        plaintiffs_str = parties.get('原告', '')
        if not plaintiffs_str or plaintiffs_str == '原告':
            return text  # 沒有具體姓名，無需修正

        # 解析原告姓名（可能有多個，用逗號分隔）
        correct_names = [name.strip() for name in plaintiffs_str.split(',')]

        # 為每個正確姓名建立可能被截斷的模式
        for correct_name in correct_names:
            if len(correct_name) < 2:
                continue  # 姓名太短，跳過

            # 姓名可能的截斷方式：
            # 例如「羅靖崴」可能被截成「羅崴」（遺漏中間字）
            # 例如「陳皆宏」可能被截成「陳宏」（遺漏中間字）

            if len(correct_name) == 3:
                # 三字名：可能遺漏中間字
                surname = correct_name[0]
                last_char = correct_name[2]

                # 匹配模式：原告[姓][名]（缺少中間字）
                truncated_pattern = f'原告{surname}{last_char}'

                # 替換所有出現的截斷姓名
                if truncated_pattern in text and f'原告{correct_name}' not in text:
                    print(f"⚠️ 發現姓名截斷：'{truncated_pattern}' → 修正為 '原告{correct_name}'")
                    text = text.replace(truncated_pattern, f'原告{correct_name}')

                # 也要處理標題中的情況（如「（一）原告羅崴之損害」）
                title_truncated_pattern = f'原告{surname}{last_char}之損害'
                title_correct_pattern = f'原告{correct_name}之損害'
                if title_truncated_pattern in text:
                    print(f"⚠️ 發現標題姓名截斷：'{title_truncated_pattern}' → 修正為 '{title_correct_pattern}'")
                    text = text.replace(title_truncated_pattern, title_correct_pattern)

                # 處理「就原告XXX部分」的格式
                conclusion_truncated = f'就原告{surname}{last_char}部分'
                conclusion_correct = f'就原告{correct_name}部分'
                if conclusion_truncated in text:
                    print(f"⚠️ 發現結論姓名截斷：'{conclusion_truncated}' → 修正為 '{conclusion_correct}'")
                    text = text.replace(conclusion_truncated, conclusion_correct)

            elif len(correct_name) == 2:
                # 二字名：較少被截斷，但還是檢查一下
                surname = correct_name[0]

                # 匹配只有姓氏的情況（極端錯誤）
                if f'原告{surname}之' in text and f'原告{correct_name}' not in text:
                    print(f"⚠️ 發現姓名嚴重截斷：'原告{surname}' → 修正為 '原告{correct_name}'")
                    # 這裡需要謹慎，避免錯誤替換（例如有多個同姓原告）
                    if len(correct_names) == 1:  # 只有一個原告才安全替換
                        text = re.sub(f'原告{surname}([之，、；。])', f'原告{correct_name}\\1', text)

        return text

    def _generate_generic_description(self, item_name: str, amount: str, plaintiff_name: str = None) -> str:
        """生成通用的項目描述"""
        plaintiff_prefix = f"原告{plaintiff_name}" if plaintiff_name else "原告"

        templates = {
            '醫療費用': f"{plaintiff_prefix}因本次事故受傷就醫，支出醫療費用{amount}元。",
            '復健費用': f"{plaintiff_prefix}因本次事故受傷需持續復健，支出復健費用{amount}元。",
            '診斷證明書費用': f"{plaintiff_prefix}購買歷次診斷證明書費用共計{amount}元。",
            '住院期間必要物品及洗頭費用': f"{plaintiff_prefix}於住院期間所支出之病房內必要物品及洗頭費用共計{amount}元。",
            '看護費用': f"{plaintiff_prefix}因本次事故受傷，需專人照護，支出看護費用{amount}元。",
            '薪資損失': f"{plaintiff_prefix}因本次事故受傷無法工作，受有薪資損失{amount}元。",
            '工作損失': f"{plaintiff_prefix}因本次事故受傷無法工作，受有工作損失{amount}元。",
            '不能工作之損失': f"{plaintiff_prefix}因本次事故受傷無法工作，受有不能工作之損失{amount}元。",
            '減少勞動能力損失': f"{plaintiff_prefix}因本次事故導致身體機能受損，減少勞動能力，受有減少勞動能力損失{amount}元。",
            '交通費用': f"{plaintiff_prefix}因本次事故就醫往返，支出交通費用{amount}元。",
            '機車修復費用': f"{plaintiff_prefix}之機車因本次事故受損，修復費用為{amount}元。",
            '車輛修復費用': f"{plaintiff_prefix}之車輛因本次事故受損，修復費用為{amount}元。",
            '車輛損壞': f"{plaintiff_prefix}之車輛因本次事故受損，修復費用為{amount}元。",
            '增加生活必要費用': f"{plaintiff_prefix}因本次事故受傷，增加生活必要費用{amount}元。",
            '精神慰撫金': f"{plaintiff_prefix}因本次事故受有傷害，身心承受相當痛苦，故請求精神慰撫金{amount}元。",
        }

        return templates.get(item_name, f"{plaintiff_prefix}因本次事故受有{item_name}{amount}元。")

    def _comprehensive_number_preprocessing(self, text: str) -> str:
        """全面預處理中文數字和特殊格式"""
        import re
        
        # 處理 X萬Y,YYY元 格式 (如：26萬4,379元)
        pattern1 = r'(\d+)萬(\d+,?\d+)元'
        def replace1(match):
            wan = int(match.group(1))
            rest = int(match.group(2).replace(',', ''))
            total = wan * 10000 + rest
            return f"{total}元"
        text = re.sub(pattern1, replace1, text)
        
        # 處理其他中文數字格式
        text = re.sub(r'(\d+)萬(\d+)千元', lambda m: f"{int(m.group(1))*10000 + int(m.group(2))*1000}元", text)
        text = re.sub(r'(\d+)萬元', lambda m: f"{int(m.group(1))*10000}元", text)
        text = re.sub(r'(\d+)千元', lambda m: f"{int(m.group(1))*1000}元", text)
        
        return text

    def _is_same_damage_type(self, context1: str, context2: str) -> bool:
        """判斷兩個上下文是否為相同的損害類型"""
        damage_types = [
            ['醫療', '治療', '就診'],
            ['看護', '照顧'],
            ['牙齒', '假牙'],
            ['慰撫', '精神', '痛苦'],
            ['交通', '車資'],
            ['工作', '收入', '薪資'],
            ['修復', '修理', '維修']
        ]
        
        # 找出每個上下文的損害類型
        type1 = None
        type2 = None
        
        for i, keywords in enumerate(damage_types):
            if any(keyword in context1 for keyword in keywords):
                type1 = i
            if any(keyword in context2 for keyword in keywords):
                type2 = i
        
        return type1 is not None and type1 == type2

    def _extract_amounts_from_damage_section(self, damage_section: str) -> list:
        """從結構化的損害項目段落中提取金額（更準確）"""
        import re

        print(f"🔍 【從損害段落提取金額】開始處理...")

        amounts = []
        lines = damage_section.split('\n')

        for line in lines:
            # 匹配損害項目標題行的金額
            # 格式1: （一）醫療費用：1,036元
            match1 = re.search(r'[（(][一二三四五六七八九十百]+[）)].*?：([\d,]+)元', line)
            if match1:
                amount_str = match1.group(1).replace(',', '')
                try:
                    amount = int(amount_str)
                    print(f"✅ 找到（中文編號）: {amount:,}元")
                    amounts.append(amount)
                    continue
                except ValueError:
                    pass

            # 格式2: 1. 醫療費用：4,862元
            match2 = re.search(r'^\d+\.\s*.*?：([\d,]+)元', line.strip())
            if match2:
                amount_str = match2.group(1).replace(',', '')
                try:
                    amount = int(amount_str)
                    print(f"✅ 找到（數字編號）: {amount:,}元")
                    amounts.append(amount)
                    continue
                except ValueError:
                    pass

        print(f"🔍 【從損害段落提取金額】共找到 {len(amounts)} 個金額")
        print(f"🔍 【從損害段落提取金額】金額列表: {amounts}")
        print(f"🔍 【從損害段落提取金額】總計: {sum(amounts):,}元")

        return amounts

    def _extract_valid_claim_amounts(self, text: str) -> list:
        """智能提取有效的求償金額（基於上下文語境）"""
        import re

        print(f"🔍 【智能金額提取】原始文本: {text[:200]}...")

        # 1. 先預處理中文數字
        processed_text = self._comprehensive_number_preprocessing(text)
        clean_text = processed_text.replace(',', '')

        # 2. 定義有效的求償關鍵詞
        valid_claim_keywords = [
            '費用', '損失', '慰撫金', '賠償', '支出', '花費',
            '醫療', '修復', '修理', '交通', '看護', '手術',
            '假牙', '復健', '治療', '工作收入', '預估', '未來', '預計', '用品',
            '薪資損害', '工資損害', '不能工作', '無法工作', '請求'  # 工資損失相關
        ]

        # 3. 定義排除的關鍵詞（非求償項目）
        # 這些關鍵詞如果出現在金額「前面」20字內，才排除
        exclude_keywords_before_amount = [
            '日薪', '年度所得', '月收入', '時薪',
            '以每', '每趟', '每月', '每日', '一日', '1日', '日1',  # 單價/日期參考
            '所得', '薪資所得', '年收入', '月薪', '底薪',  # 薪資參考數據
            '行情為', '收費行情', '單價', '收費',  # 價格行情
        ]

        # 這些關鍵詞出現在上下文任何位置都排除
        exclude_keywords_anywhere = [
            '學歷', '畢業', '名下', '動產',
            '包括', '其中', '包含',  # 細項分解關鍵詞
            '經查', '查明', '經審理'  # 判決書用語
        ]

        # 4. 定義需要更寬鬆判斷的證據關鍵詞（這些詞出現時不一定要排除）
        # 只有在沒有求償關鍵詞的情況下才排除
        evidence_keywords = ['此有', '可證', '為證', '收據', '發票', '證明']

        amounts = []
        lines = clean_text.split('\n')

        for line in lines:
            # 找出該行中的所有金額
            line_amounts = re.findall(r'(\d+)\s*元', line)

            for amt_str in line_amounts:
                try:
                    amount = int(amt_str)
                    if amount < 100:  # 跳過小額（可能是編號等）
                        continue

                    # 檢查金額周圍的上下文
                    # 找到金額在原文中的位置
                    amount_pos = line.find(amt_str + '元')
                    if amount_pos == -1:
                        continue

                    # 提取金額前後的上下文（前50字，後50字）
                    start = max(0, amount_pos - 50)
                    end = min(len(line), amount_pos + 50)
                    context = line[start:end]

                    # 提取金額前20字（用於檢查緊鄰金額的排除詞）
                    before_start = max(0, amount_pos - 20)
                    context_before = line[before_start:amount_pos]

                    # 先檢查是否包含有效求償關鍵詞
                    is_valid_claim = any(keyword in context for keyword in valid_claim_keywords)

                    if is_valid_claim:
                        # 如果是有效求償項目，再檢查是否需要排除
                        # 1. 檢查金額前是否有排除關鍵詞（如"月薪7萬元"）
                        should_exclude_before = any(keyword in context_before for keyword in exclude_keywords_before_amount)
                        # 2. 檢查整個上下文是否有絕對排除詞
                        should_exclude_anywhere = any(keyword in context for keyword in exclude_keywords_anywhere)

                        should_exclude = should_exclude_before or should_exclude_anywhere

                        # 特殊處理：如果金額前有「支出」「共」「合計」「總計」等總額關鍵詞，不排除
                        total_keywords = ['支出', '共', '合計', '總計', '一共']
                        has_total_keyword = any(keyword in context_before for keyword in total_keywords)
                        if has_total_keyword:
                            should_exclude = False

                        if should_exclude:
                            print(f"🔍 【排除】{amount:,}元 - 包含排除關鍵詞: {context[:50]}...")
                        else:
                            print(f"✅ 【有效】{amount:,}元 - 上下文: {context[:50]}...")
                            amounts.append(amount)
                    else:
                        # 沒有明確求償關鍵詞時，檢查是否只是證據引用
                        has_evidence_only = any(keyword in context for keyword in evidence_keywords)
                        if has_evidence_only:
                            print(f"🔍 【排除】{amount:,}元 - 僅為證據引用: {context[:50]}...")
                        else:
                            print(f"🔍 【跳過】{amount:,}元 - 無明確求償關鍵詞: {context[:50]}...")

                except ValueError:
                    continue

        # 4. 改進的去重邏輯（按項目類型分組）
        damage_items = {}  # 按類型分組：{類型: [金額列表]}
        
        for line in clean_text.split('\n'):
            # 識別損害項目標題行（如：（一）醫療費用38,073元 或 1. 醫療費用38,073元）
            if (re.match(r'^[（][一二三四五六七八九十][）]', line.strip()) or 
                re.match(r'^[㈠㈡㈢㈣㈤㈥㈦㈧㈨㈩]', line.strip()) or 
                re.match(r'^\d+\.\s*[^\d]*\d+元', line.strip())):
                line_amounts = re.findall(r'(\d+)\s*元', line)
                for amt_str in line_amounts:
                    try:
                        amount = int(amt_str)
                        if amount >= 100:  # 排除小額
                            # 判斷損害類型
                            damage_type = "其他"
                            if '預估醫療' in line or '未來醫療' in line or '預計醫療' in line:
                                damage_type = "預估醫療費用"
                            elif '醫療用品' in line:
                                damage_type = "醫療用品費用"
                            elif '醫療' in line:
                                damage_type = "醫療費用"
                            elif '看護' in line:
                                damage_type = "看護費用"
                            elif '牙齒' in line or '假牙' in line:
                                damage_type = "牙齒損害"
                            elif '慰撫' in line or '精神' in line:
                                damage_type = "精神慰撫金"
                            elif '交通' in line:
                                damage_type = "交通費用"
                            elif '車輛' in line or '機車' in line or '修復' in line or '修理' in line or '維修' in line:
                                damage_type = "車輛修復費用"
                            elif '無法工作' in line or '工作損失' in line:
                                damage_type = "無法工作損失"
                            elif '工作' in line or '收入' in line or '損失' in line:
                                damage_type = "工作損失"
                            
                            if damage_type not in damage_items:
                                damage_items[damage_type] = []
                            damage_items[damage_type].append(amount)
                            print(f"🔍 【確認項目】{damage_type}: {amount:,}元")
                    except ValueError:
                        continue
        
        # 每種損害類型只取一個金額（通常標題行的金額是正確的）
        final_amounts = []
        for damage_type, amounts_list in damage_items.items():
            if amounts_list:
                # 取該類型的第一個金額（標題行）
                final_amounts.append(amounts_list[0])
                print(f"✅ 【採用】{damage_type}: {amounts_list[0]:,}元")

        # 如果沒有找到標題行格式，回退使用context-based提取的結果
        if not final_amounts and amounts:
            print(f"⚠️  未找到標題行格式，使用context-based提取結果進行去重")
            # 簡單去重：移除重複值，保持順序
            seen = set()
            for amt in amounts:
                if amt not in seen:
                    final_amounts.append(amt)
                    seen.add(amt)
                    print(f"✅ 【採用】{amt:,}元")

        print(f"🔍 【智能金額提取】去重後有效金額: {final_amounts}")
        print(f"🔍 【智能金額提取】最終總計: {sum(final_amounts):,}元")

        return final_amounts

    def _extract_damage_items_from_text(self, text: str) -> Dict[str, List[Dict]]:
        """從文本中精確提取損害項目"""
        # 按原告分組
        plaintiff_damages = {}
        
        # 分句處理
        sentences = re.split(r'[。]', text)
        
        for sentence in sentences:
            # 識別原告
            plaintiff_match = re.search(r'原告([^，。；、\s]{2,4})', sentence)
            if not plaintiff_match:
                continue
                
            plaintiff = plaintiff_match.group(1)
            if plaintiff not in plaintiff_damages:
                plaintiff_damages[plaintiff] = []
            
            # 精確匹配各種損害類型
            # 醫療費用
            if '醫療費用' in sentence:
                amount_match = re.search(r'醫療費用\s*(\d+(?:,\d{3})*)\s*元', sentence)
                if amount_match:
                    plaintiff_damages[plaintiff].append({
                        'name': '醫療費用',
                        'amount': int(amount_match.group(1).replace(',', '')),
                        'description': f'原告{plaintiff}因本次事故受傷就醫之醫療費用'
                    })
            
            # 交通費
            if '交通費' in sentence:
                amount_match = re.search(r'交通費\s*(\d+(?:,\d{3})*)\s*元', sentence)
                if amount_match:
                    plaintiff_damages[plaintiff].append({
                        'name': '交通費用',
                        'amount': int(amount_match.group(1).replace(',', '')),
                        'description': f'原告{plaintiff}因本次事故所生之交通費用'
                    })
            
            # 工作損失
            if any(keyword in sentence for keyword in ['工資損失', '不能工作', '無法工作']):
                amount_match = re.search(r'(?:損失|請求)\s*(\d+(?:,\d{3})*)\s*元', sentence)
                if amount_match:
                    plaintiff_damages[plaintiff].append({
                        'name': '工作損失',
                        'amount': int(amount_match.group(1).replace(',', '')),
                        'description': f'原告{plaintiff}因本次事故無法工作之收入損失'
                    })
            
            # 精神慰撫金
            if '慰撫金' in sentence:
                amount_match = re.search(r'慰撫金\s*(\d+(?:,\d{3})*)\s*元', sentence)
                if amount_match:
                    plaintiff_damages[plaintiff].append({
                        'name': '精神慰撫金',
                        'amount': int(amount_match.group(1).replace(',', '')),
                        'description': f'原告{plaintiff}因本次事故所受精神痛苦之慰撫金'
                    })
            
            # 車輛貶值
            if any(keyword in sentence for keyword in ['貶值', '貶損', '價值減損']):
                amount_match = re.search(r'(?:貶損|減損)\s*(\d+(?:,\d{3})*)\s*元', sentence)
                if amount_match:
                    plaintiff_damages[plaintiff].append({
                        'name': '車輛貶值損失',
                        'amount': int(amount_match.group(1).replace(',', '')),
                        'description': '系爭車輛因本次事故之價值減損'
                    })
            
            # 鑑定費
            if '鑑定費' in sentence:
                amount_match = re.search(r'鑑定費\s*(\d+(?:,\d{3})*)\s*元', sentence)
                if amount_match:
                    plaintiff_damages[plaintiff].append({
                        'name': '鑑定費用',
                        'amount': int(amount_match.group(1).replace(',', '')),
                        'description': '車輛損害鑑定費用'
                    })
        
        return plaintiff_damages
    
    def _format_damage_items(self, damage_items: Dict[str, List[Dict]]) -> str:
        """格式化損害項目"""
        if not damage_items:
            return ""
        
        result = ""
        for idx, (plaintiff, damages) in enumerate(damage_items.items()):
            chinese_num = self._chinese_num(idx + 1)
            result += f"\n（{chinese_num}）原告{plaintiff}之損害：\n"
            
            for i, damage in enumerate(damages, 1):
                result += f"{i}. {damage['name']}：{damage['amount']:,}元\n"
                result += f"   說明：{damage['description']}\n"
            
            # 小計
            subtotal = sum(d['amount'] for d in damages)
            result += f"\n小計：{subtotal:,}元\n"
        
        # 總計
        total = sum(sum(d['amount'] for d in damages) for damages in damage_items.values())
        result += f"\n損害總計：新台幣{total:,}元整"
        
        return result

# ===== 主要互動功能 =====

def interactive_generate_lawsuit():
    """互動式起訴狀生成（恢復多行輸入版本）"""
    print("=" * 80)
    print("🏛️  車禍起訴狀生成器 - 混合版本（整合結構化金額處理）")
    print("=" * 80)
    print("👋 歡迎使用！我會為您生成專業的起訴狀，包含：")
    print("   📄 相似案例檢索")
    print("   ⚖️ 適用法條分析")
    print("   📋 完整起訴狀生成")
    print("💡 支援結構化金額處理，自動修正計算錯誤")
    print()
    
    print("📝 使用方法：")
    print("   1. 請一次性輸入完整的三段內容")
    print("   2. 可以多行輸入，換行繼續")
    print("   3. 輸入完成後輸入 'END' 確認")
    print("   4. 輸入 'quit' 可退出程式")
    print()
    
    # 初始化生成器
    generator = HybridCoTGenerator()
    
    print("📝 請輸入完整的車禍案件資料：")
    print("📋 請包含以下三個部分：")
    print("   一、事故發生緣由：[詳述車禍經過]")
    print("   二、原告受傷情形：[描述傷勢]")
    print("   三、請求賠償的事實根據：[列出損害項目和金額]")
    print()
    print("💡 提示：可以換行輸入，完成後輸入 'END' 確認")
    print("=" * 60)
    print("🎯 請開始輸入（完成後輸入 'END' 或 'end' 確認）：")
    
    # 多行輸入模式
    user_input_lines = []
    while True:
        try:
            line = input()
            if line.strip().upper() in ['END', 'QUIT', 'EXIT', '退出']:
                if line.strip().upper() == 'QUIT' or line.strip().upper() == 'EXIT' or line.strip() == '退出':
                    print("👋 感謝使用，再見！")
                    return
                break
            user_input_lines.append(line)
        except KeyboardInterrupt:
            print("\n👋 用戶中斷，程序退出")
            return
        except EOFError:
            break
    
    user_query = '\n'.join(user_input_lines).strip()
    
    if not user_query:
        print("⚠️ 請輸入有效的內容")
        return
    
    print("🔄 正在處理...")
    
    try:
        # 分段提取資訊
        sections = extract_sections(user_query)
        
        # 提取當事人
        parties = extract_parties(user_query)
        
        # 案件分類和檢索相似案例
        accident_facts = sections.get("accident_facts", user_query)
        case_type = determine_case_type(accident_facts, parties)
        
        print("✅ LLM服務正常")
        print()
        
        # 相似案例檢索（詳細模式）
        similar_cases = []
        if FULL_MODE:
            print("🔍 檢索相似案例...")
            # 使用固定參數進行檢索
            k_final = 3
            initial_retrieve_count = 15
            use_multi_stage = True
            
            print(f"🔧 檢索策略: 目標{k_final}個案例 → 檢索{initial_retrieve_count}個段落")
            
            try:
                # 詳細執行檢索
                query_vector = embed(accident_facts)
                if query_vector:
                    hits = es_search(query_vector, case_type, top_k=initial_retrieve_count, label="Facts", quiet=False)
                    if hits:
                        print(f"🔍 ES原始結果: 找到{len(hits)}個段落")
                        
                        # 多階段模式
                        candidate_case_ids = []
                        for hit in hits:
                            case_id = hit['_source'].get('case_id')
                            if case_id and case_id not in candidate_case_ids:
                                candidate_case_ids.append(case_id)
                        
                        print(f"🔄 去重後結果: {len(candidate_case_ids)}個唯一案例")
                        final_case_ids = candidate_case_ids[:k_final]
                        print(f"📌 最終選取: {len(final_case_ids)}個案例 {final_case_ids}")
                        
                        if candidate_case_ids:
                            reranked_case_ids = rerank_case_ids_by_paragraphs(
                                accident_facts, 
                                candidate_case_ids[:k_final*2],
                                label="Facts",
                                quiet=False
                            )
                            final_case_ids = reranked_case_ids[:k_final]
                            print(f"📘 Rerank後最終順序: {final_case_ids}")
                            
                            similar_cases = get_complete_cases_content(final_case_ids)
                            
                            # 顯示詳細案例分析
                            print()
                            print(f"📋 詳細案例分析 (僅顯示前 {len(similar_cases)} 個最相關案例):")
                            print("=" * 80)
                            print()
                            
                            for i, (case_content, case_id) in enumerate(zip(similar_cases, final_case_ids)):
                                # 從hits中找到對應的分數
                                score = 0.0
                                for hit in hits:
                                    if hit['_source'].get('case_id') == case_id:
                                        score = hit['_score']
                                        break
                                
                                print(f"📄 相似案例 {i+1}: Case ID {case_id}")
                                print(f"🎯 ES相似度分數: {score:.4f}")
                                print("-" * 50)
                                case_preview = case_content[:500] + "..." if len(case_content) > 500 else case_content
                                print(case_preview)
                                print()
                                if i < len(similar_cases) - 1:
                                    print()
                        else:
                            # 簡單模式
                            if hits:
                                first_hit = hits[0]['_source']
                                content_fields = ['original_text', 'content', 'text', 'facts_content', 'chunk_content', 'body']
                                content_field = None
                                for field in content_fields:
                                    if field in first_hit and first_hit[field]:
                                        content_field = field
                                        break
                                
                                if content_field:
                                    similar_cases = [hit['_source'].get(content_field, '') for hit in hits[:k_final] if hit['_source'].get(content_field)]
                                else:
                                    # Fallback
                                    similar_cases = []
                                    for hit in hits[:k_final]:
                                        all_text = " ".join([str(v) for k, v in hit['_source'].items() 
                                                           if isinstance(v, str) and len(v) > 50])
                                        if all_text:
                                            similar_cases.append(all_text)
            except Exception as e:
                similar_cases = []
        
        # ===== 選擇生成模式 =====
        print(f"\n🎯 請選擇生成模式：")
        print("1. 標準混合模式（事實+法條+損害：標準方式，結論：CoT方式）")
        print("2. 雙輪CoT模式（生成→檢查→修正的完整驗證流程）")
        
        while True:
            try:
                mode_choice = input("請選擇模式 (1 或 2): ").strip()
                if mode_choice in ['1', '2']:
                    break
                print("請輸入 1 或 2")
            except KeyboardInterrupt:
                print("\n👋 用戶中斷，程序退出")
                return
        
        if mode_choice == '2':
            # ===== 雙輪CoT模式 =====
            print(f"\n🔄 開始雙輪CoT模式生成...")
            print("📝 輪1：生成完整起訴狀")
            print("🔍 輪2：檢查驗證並修正")
            print()
            
            # 使用雙輪CoT生成完整起訴狀
            final_result = generator.generate_complete_dual_round_cot(
                accident_facts,
                sections.get("compensation_facts", user_query),
                parties,
                similar_cases,
                final_case_ids if 'final_case_ids' in locals() else []
            )
            
            # 輸出完整結果
            print("=" * 60)
            print("⚖️ 適用法條")
            print("=" * 60)
            for i, law in enumerate(final_result.get('applicable_laws', []), 1):
                print(f"{i}. {law}")
            
            print("\n" + "=" * 60)
            print("📋 生成的起訴狀")
            print("=" * 60)
            print(final_result.get('lawsuit', ''))
            
            print("\n" + "=" * 60)
            print("✅ 起訴狀生成完成!")
            return
        
        # ===== 標準混合模式生成 =====
        print(f"\n🎯 開始標準混合模式生成...")
        print("📝 事實 + 法條 + 損害：標準方式")
        print("🧠 結論：CoT方式（計算總金額）")
        print()
        
        # 生成事實段落
        print("📝 生成事實段落...")
        facts = generator.generate_standard_facts(accident_facts, similar_cases)
        print("✅ 事實段落生成完成")
        
        # 生成法律依據
        print("⚖️ 生成法律依據...")
        
        # 統計相似案例使用的法條
        if similar_cases and 'final_case_ids' in locals():
            try:
                print("📊 分析相似案例使用的法條...")
                similar_laws_stats = get_similar_cases_laws_stats(final_case_ids)
                if similar_laws_stats:
                    print("📋 相似案例常用法條統計:")
                    for law_name, count in similar_laws_stats[:5]:  # 顯示前5個最常用的
                        print(f"   • {law_name}: {count}次")
                    print()
            except Exception as e:
                print(f"⚠️ 法條統計分析失敗: {e}")
        
        laws = generator.generate_standard_laws(
            sections.get("accident_facts", user_query),
            sections.get("injuries", ""),
            parties,
            sections.get("compensation_facts", "")
        )
        print("✅ 法律依據生成完成")
        
        # 生成損害賠償
        print("💰 生成損害賠償...")
        compensation_text = sections.get("compensation_facts", user_query)
        damages = generator.generate_smart_compensation(
            sections.get("injuries", ""),
            compensation_text, 
            parties
        )
        print("✅ 損害賠償生成完成")
        
        # 生成CoT結論
        print("🧠 生成CoT結論（含總金額計算）...")
        conclusion = generator.generate_cot_conclusion_with_structured_analysis(
            sections.get("accident_facts", user_query),
            compensation_text,
            parties,
            damages  # 傳入損害項目以便計算編號
        )
        print("✅ CoT結論生成完成")
        print()
        print("✅ 所有生成步驟完成！")
        
        # 提取適用法條
        applicable_laws = determine_applicable_laws(
            sections.get("accident_facts", user_query),
            sections.get("injuries", ""),
            sections.get("compensation_facts", ""),
            parties
        )
        
        # ===== 輸出核心結果 =====
        print("\n" + "=" * 60)
        print("⚖️ 適用法條")
        print("=" * 60)
        
        for i, law in enumerate(applicable_laws, 1):
            print(f"{i}. {law}")
        
        print("\n" + "=" * 60)
        print("📋 生成的起訴狀")
        print("=" * 60)
        
        print(f"\n{facts}")
        print(f"\n{laws}")
        print(f"\n{damages}")
        print(f"\n{conclusion}")
        
        print("\n" + "=" * 60)
        print("✅ 起訴狀生成完成!")
        
    except Exception as e:
        import traceback
        print(f"❌ 生成過程中發生錯誤：{str(e)}")
        print(f"❌ 詳細錯誤信息：{traceback.format_exc()}")
        print("請檢查輸入格式或聯繫系統管理員")

def batch_from_excel(excel_path: str, output_col: str = "D", sheet_index: int = 0):
    """批次處理 Excel：讀取 C 欄律師輸入，生成起訴狀寫回指定欄位"""
    import openpyxl

    print(f"📂 開啟 Excel：{excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.worksheets[sheet_index]
    print(f"📋 工作表：{ws.title}，共 {ws.max_row - 1} 列資料")
    print(f"📊 檢索模式：{'完整模式' if FULL_MODE else '簡化模式'}")
    print("=" * 60)

    generator = HybridCoTGenerator()

    # 找出 C 欄和輸出欄的欄號
    input_col_idx = 3   # C 欄
    output_col_idx = openpyxl.utils.column_index_from_string(output_col)

    success_count = 0
    skip_count = 0
    error_count = 0

    for row_idx in range(2, ws.max_row + 1):
        lawyer_input = ws.cell(row=row_idx, column=input_col_idx).value
        existing_output = ws.cell(row=row_idx, column=output_col_idx).value

        if not lawyer_input or not str(lawyer_input).strip():
            skip_count += 1
            continue

        if existing_output and str(existing_output).strip():
            print(f"⏭️  第 {row_idx} 列（case_id={ws.cell(row=row_idx, column=2).value}）已有結果，跳過")
            skip_count += 1
            continue

        case_id = ws.cell(row=row_idx, column=2).value
        print(f"\n{'='*60}")
        print(f"🔄 處理第 {row_idx} 列 / case_id={case_id} ({success_count + error_count + 1}/{ws.max_row - 1})")
        print("=" * 60)

        try:
            user_query = str(lawyer_input).strip()

            # 分段解析
            sections = extract_sections(user_query)
            parties = extract_parties(user_query)
            accident_facts = sections.get("accident_facts", user_query)
            case_type = determine_case_type(accident_facts, parties)

            # 相似案例檢索
            similar_cases = []
            final_case_ids = []
            if FULL_MODE:
                try:
                    query_vector = embed(accident_facts)
                    if query_vector:
                        hits = es_search(query_vector, case_type, top_k=15, label="Facts", quiet=True)
                        if hits:
                            candidate_ids = []
                            for hit in hits:
                                cid = hit['_source'].get('case_id')
                                if cid and cid not in candidate_ids:
                                    candidate_ids.append(cid)
                            reranked = rerank_case_ids_by_paragraphs(
                                accident_facts, candidate_ids[:6], label="Facts", quiet=True
                            )
                            final_case_ids = reranked[:3]
                            similar_cases = get_complete_cases_content(final_case_ids)
                            print(f"📌 檢索到相似案例：{final_case_ids}")
                except Exception as e:
                    print(f"⚠️ 相似案例檢索失敗：{e}")

            # 生成各段落（標準混合模式）
            print("📝 生成事實段落...")
            facts = generator.generate_standard_facts(accident_facts, similar_cases)

            print("⚖️ 生成法律依據...")
            laws = generator.generate_standard_laws(
                sections.get("accident_facts", user_query),
                sections.get("injuries", ""),
                parties,
                sections.get("compensation_facts", "")
            )

            print("💰 生成損害賠償...")
            damages = generator.generate_smart_compensation(
                sections.get("injuries", ""),
                sections.get("compensation_facts", user_query),
                parties
            )

            print("🧠 生成CoT結論...")
            conclusion = generator.generate_cot_conclusion_with_structured_analysis(
                accident_facts,
                sections.get("compensation_facts", user_query),
                parties,
                damages
            )

            # 組合結果
            result = f"{facts}\n\n{laws}\n\n{damages}\n\n{conclusion}"
            ws.cell(row=row_idx, column=output_col_idx).value = result

            # 每筆完成後立即儲存，避免中斷遺失
            wb.save(excel_path)
            success_count += 1
            print(f"✅ 第 {row_idx} 列完成，已儲存")

        except Exception as e:
            import traceback
            print(f"❌ 第 {row_idx} 列發生錯誤：{e}")
            print(traceback.format_exc())
            error_count += 1
            # 發生錯誤也記錄到欄位，方便追蹤
            ws.cell(row=row_idx, column=output_col_idx).value = f"[ERROR] {str(e)}"
            wb.save(excel_path)

    print(f"\n{'='*60}")
    print(f"🏁 批次處理完成")
    print(f"   ✅ 成功：{success_count} 筆")
    print(f"   ⏭️  跳過：{skip_count} 筆")
    print(f"   ❌ 錯誤：{error_count} 筆")
    print(f"   💾 結果已儲存至：{excel_path}")
    print("=" * 60)


def main():
    """主程序入口"""
    parser = argparse.ArgumentParser(description="車禍起訴狀生成器")
    parser.add_argument(
        "--batch",
        metavar="EXCEL_PATH",
        help="批次模式：指定 Excel 檔案路徑，讀取 C 欄律師輸入，結果寫回 D 欄"
    )
    parser.add_argument(
        "--output-col",
        default="D",
        metavar="COL",
        help="批次模式輸出欄位（預設：D）"
    )
    parser.add_argument(
        "--sheet",
        type=int,
        default=0,
        metavar="N",
        help="批次模式使用第幾張工作表，從 0 開始（預設：0）"
    )
    args = parser.parse_args()

    try:
        print("🔧 檢查系統依賴...")
        print(f"📊 檢索模式：{'完整模式' if FULL_MODE else '簡化模式'}")
        print(f"🏗️ 結構化處理器：{'可用' if STRUCTURED_PROCESSOR_AVAILABLE else '不可用'}")
        print(f"📏 基本標準化器：{'可用' if BASIC_STANDARDIZER_AVAILABLE else '不可用'}")

        if args.batch:
            batch_from_excel(args.batch, output_col=args.output_col, sheet_index=args.sheet)
        else:
            interactive_generate_lawsuit()

    except KeyboardInterrupt:
        print("\n\n👋 用戶中斷，程序退出")
    except Exception as e:
        print(f"\n❌ 程序執行錯誤：{str(e)}")

if __name__ == "__main__":
    main()