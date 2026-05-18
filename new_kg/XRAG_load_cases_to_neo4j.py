"""
XRAG case loader for Neo4j Aura.

Purpose:
- Load the 6057 fixed case nodes into Neo4j once
- Keep case-level features stable across all later experiments
- Avoid the old trie/hash ontology script, which is no longer the core path

Data sources:
- phase1_boolean_matrix_v1.jsonl: fixed 16-dim boolean encoding matrix
- 6057律師輸入.xlsx: lawyer_input text
- 6057起訴書_全.xlsx: full complaint text

Node model:
  (:Case {case_id})

Selected properties:
  - lawyer_input
  - complaint_text
  - fact_text / laws_text / compensation_text / conclusion_text
  - 16 boolean matrix fields
  - boolean_matrix_json / tensor_json / review_flags_json
  - legacy_hints_json

Usage:
  python XRAG_load_cases_to_neo4j.py
  python XRAG_load_cases_to_neo4j.py --wipe
"""

from __future__ import annotations

import argparse
import json
import os
from pathlib import Path
from typing import Dict, Iterable, List

from neo4j import GraphDatabase
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
BOOLEAN_FILE = BASE_DIR / "phase1_boolean_matrix_v1.jsonl"
LAWYER_INPUT_FILE = Path("/home/aru/AI_LAW/09_輸入輸出資料/6057律師輸入.xlsx")
COMPLAINT_FILE = Path("/home/aru/AI_LAW/6057起訴書_全.xlsx")
ENV_FILE = BASE_DIR / ".env"

BATCH_SIZE = 100


def load_env_file(path: Path) -> Dict[str, str]:
    env: Dict[str, str] = {}
    if not path.exists():
        return env
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip()
    return env


def get_config() -> Dict[str, str]:
    file_env = load_env_file(ENV_FILE)
    keys = ["NEO4J_URI", "NEO4J_USERNAME", "NEO4J_PASSWORD", "NEO4J_DATABASE"]
    cfg = {}
    for key in keys:
        cfg[key] = os.environ.get(key) or file_env.get(key, "")
    missing = [k for k, v in cfg.items() if not v]
    if missing:
        raise RuntimeError(f"Missing Neo4j config: {', '.join(missing)}")
    return cfg


def read_xlsx_column_map(path: Path, required: List[str]) -> Dict[str, Dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header_row = next(rows)
    header = [str(c).strip() if c is not None else "" for c in header_row]
    index = {name: idx for idx, name in enumerate(header)}

    missing = [col for col in required if col not in index]
    if missing:
        raise RuntimeError(f"{path.name} missing required columns: {missing}")

    result: Dict[str, Dict[str, str]] = {}
    for row in rows:
        case_id = row[index["case_id"]]
        if case_id is None:
            continue
        cid = str(case_id).strip()
        result[cid] = {}
        for col in required:
            if col == "case_id":
                continue
            value = row[index[col]]
            result[cid][col] = "" if value is None else str(value)
    return result


def load_lawyer_inputs() -> Dict[str, str]:
    data = read_xlsx_column_map(LAWYER_INPUT_FILE, ["case_id", "律師輸入"])
    return {cid: row["律師輸入"] for cid, row in data.items()}


def load_full_complaints() -> Dict[str, str]:
    data = read_xlsx_column_map(COMPLAINT_FILE, ["case_id", "起訴書"])
    return {cid: row["起訴書"] for cid, row in data.items()}


def iter_records(path: Path) -> Iterable[dict]:
    with path.open(encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if line:
                yield json.loads(line)


def build_case_rows() -> List[dict]:
    lawyer_inputs = load_lawyer_inputs()
    full_complaints = load_full_complaints()

    rows: List[dict] = []
    for rec in iter_records(BOOLEAN_FILE):
        cid = str(rec["case_id"])
        matrix = rec.get("boolean_matrix", {})
        litigants = matrix.get("litigants", {})
        fact = matrix.get("fact", {})
        injury = matrix.get("injury", {})
        compensation = matrix.get("compensation", {})

        complaint_text = full_complaints.get(cid, "")
        if not complaint_text:
            complaint_text = "\n".join(
                part for part in [
                    rec.get("fact_text", ""),
                    rec.get("laws_text", ""),
                    rec.get("compensation_text", ""),
                    rec.get("conclusion_text", ""),
                ]
                if part
            )

        rows.append({
            "case_id": cid,
            "lawyer_input": lawyer_inputs.get(cid, ""),
            "complaint_text": complaint_text,
            "fact_text": rec.get("fact_text", ""),
            "laws_text": rec.get("laws_text", ""),
            "compensation_text": rec.get("compensation_text", ""),
            "conclusion_text": rec.get("conclusion_text", ""),
            "single_plaintiff": int(litigants.get("single_plaintiff", 0)),
            "multiple_plaintiffs": int(litigants.get("multiple_plaintiffs", 0)),
            "single_defendant": int(litigants.get("single_defendant", 0)),
            "multiple_defendants": int(litigants.get("multiple_defendants", 0)),
            "negligence": int(fact.get("negligence", 0)),
            "gross_negligence": int(fact.get("gross_negligence", 0)),
            "joint_liability": int(fact.get("joint_liability", 0)),
            "prior_criminal": int(fact.get("prior_criminal", 0)),
            "head_neck": int(injury.get("head_neck", 0)),
            "trunk": int(injury.get("trunk", 0)),
            "extremities": int(injury.get("extremities", 0)),
            "psych_other": int(injury.get("psych_other", 0)),
            "medical_rehab": int(compensation.get("medical_rehab", 0)),
            "lost_income": int(compensation.get("lost_income", 0)),
            "non_pecuniary": int(compensation.get("non_pecuniary", 0)),
            "care_other": int(compensation.get("care_other", 0)),
            "boolean_matrix_json": json.dumps(matrix, ensure_ascii=False),
            "tensor_json": json.dumps(rec.get("tensor", []), ensure_ascii=False),
            "review_flags_json": json.dumps(rec.get("review_flags", []), ensure_ascii=False),
            "legacy_hints_json": json.dumps(rec.get("legacy_hints", {}), ensure_ascii=False),
        })

    return rows


def ensure_schema(driver, database: str) -> None:
    statements = [
        "CREATE CONSTRAINT case_id_unique IF NOT EXISTS FOR (c:Case) REQUIRE c.case_id IS UNIQUE",
        "CREATE INDEX case_litigant_flags IF NOT EXISTS FOR (c:Case) ON (c.single_plaintiff, c.multiple_plaintiffs, c.single_defendant, c.multiple_defendants)",
        "CREATE INDEX case_fact_flags IF NOT EXISTS FOR (c:Case) ON (c.negligence, c.gross_negligence, c.joint_liability, c.prior_criminal)",
    ]
    with driver.session(database=database) as session:
        for cypher in statements:
            session.run(cypher)


def wipe_cases(driver, database: str) -> None:
    with driver.session(database=database) as session:
        session.run("MATCH (c:Case) DETACH DELETE c")


def load_cases(driver, database: str, rows: List[dict], batch_size: int) -> None:
    query = """
    UNWIND $rows AS row
    MERGE (c:Case {case_id: row.case_id})
    SET c.lawyer_input = row.lawyer_input,
        c.complaint_text = row.complaint_text,
        c.fact_text = row.fact_text,
        c.laws_text = row.laws_text,
        c.compensation_text = row.compensation_text,
        c.conclusion_text = row.conclusion_text,
        c.single_plaintiff = row.single_plaintiff,
        c.multiple_plaintiffs = row.multiple_plaintiffs,
        c.single_defendant = row.single_defendant,
        c.multiple_defendants = row.multiple_defendants,
        c.negligence = row.negligence,
        c.gross_negligence = row.gross_negligence,
        c.joint_liability = row.joint_liability,
        c.prior_criminal = row.prior_criminal,
        c.head_neck = row.head_neck,
        c.trunk = row.trunk,
        c.extremities = row.extremities,
        c.psych_other = row.psych_other,
        c.medical_rehab = row.medical_rehab,
        c.lost_income = row.lost_income,
        c.non_pecuniary = row.non_pecuniary,
        c.care_other = row.care_other,
        c.boolean_matrix_json = row.boolean_matrix_json,
        c.tensor_json = row.tensor_json,
        c.review_flags_json = row.review_flags_json,
        c.legacy_hints_json = row.legacy_hints_json
    """

    with driver.session(database=database) as session:
        for start in range(0, len(rows), batch_size):
            batch = rows[start:start + batch_size]
            session.run(query, rows=batch)
            print(f"Loaded {start + len(batch)}/{len(rows)} cases")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--wipe", action="store_true", help="Delete existing :Case nodes before loading")
    parser.add_argument("--batch-size", type=int, default=BATCH_SIZE)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    cfg = get_config()

    rows = build_case_rows()
    print(f"Prepared {len(rows)} case rows from fixed feature data")

    driver = GraphDatabase.driver(
        cfg["NEO4J_URI"],
        auth=(cfg["NEO4J_USERNAME"], cfg["NEO4J_PASSWORD"]),
    )

    ensure_schema(driver, cfg["NEO4J_DATABASE"])
    if args.wipe:
        print("Deleting existing :Case nodes ...")
        wipe_cases(driver, cfg["NEO4J_DATABASE"])

    load_cases(driver, cfg["NEO4J_DATABASE"], rows, args.batch_size)
    driver.close()
    print("Done.")


if __name__ == "__main__":
    main()
